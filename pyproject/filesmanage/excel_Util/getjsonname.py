#提取json的文件名+json文档里"文档名称"各一列

#改名
# "城市公共设施规划规范GB504422008",
# "电磁屏蔽室工程施工及质量验收规范GBT511032015",
# "沉井与气压沉箱施工规范"
# 成：
# 《城市公共设施规划规范》GB 50442-2008
# 《电磁屏蔽室工程施工及质量验收规范》GB T51103-2015
# 《沉井与气压沉箱施工规范》
import re
import os
import sys
from pathlib import Path
import json
import shutil
import openpyxl

# 获取当前文件的父目录
parent_dir = str(Path(__file__).resolve().parent.parent)
# 将父目录添加到sys.path
sys.path.append(parent_dir)

from filesfunction import opfiles

# 定义文件夹路径
source_folder, parent_folder = opfiles.OpFiles.select_folder()

filenames = []
jsonnames = []
icount = 0
icount2 = 0
# 遍历源文件夹中的所有文件
for filename in os.listdir(source_folder):
    icount = icount + 1  
    if filename.endswith('.json') or filename.endswith('.Json'):
        file_path = os.path.join(source_folder, filename)
        #jsonnames.append(filename)
        # 打开并读取JSON文件
        with open(file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
  
        try:
            for entry in data:
                icount2 = icount2 + 1    
                if "文档名称" in entry:
                    name = entry["文档名称"]
                    filenames.append(name)
                    jsonnames.append(filename)
 
                    break
                else:
                    print("entry does not have a '版本' key")
            

        except json.JSONDecodeError:
            print(f'{filename} 不是有效的 JSON 文件')
    if not icount2 == icount:
        pass

# 创建一个新的Excel工作簿和工作表
workbook = openpyxl.Workbook()
sheet = workbook.active

row = 0
for item in jsonnames:
    row += 1
    sheet.cell(row=row, column=1, value=item)
   

row = 0
for item in filenames:
    row += 1
    sheet.cell(row=row, column=3, value=item)
  
excel_file_name = source_folder.split("/")[-1] + "_提取文件名.xlsx"
excelfolder_path = os.path.normpath(os.path.join(parent_folder, excel_file_name))
workbook.save(excelfolder_path)