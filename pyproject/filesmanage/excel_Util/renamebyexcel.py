#改文件名用excel提供的信息
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import os
import sys
from pathlib import Path

# 获取当前文件的父目录
parent_dir = str(Path(__file__).resolve().parent.parent)
# 将父目录添加到sys.path
sys.path.append(parent_dir)

from filesfunction import opfiles

# 创建 Tkinter 主窗口（通常隐藏）
root = tk.Tk()
root.withdraw()

# 打开文件选择对话框
file_path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel文件', '*.xlsx')])

source_folder, parent_folder = opfiles.OpFiles.select_folder()

newfolder = source_folder.split("/")[-1] + "_newname"
newfolder_path = os.path.normpath(os.path.join(parent_folder, newfolder))
try:
    os.makedirs(newfolder_path)
except FileExistsError:
    pass

# 检查用户是否选择了文件
if file_path:
    # 加载选定的 Excel 文件
    workbook = load_workbook(filename=file_path)
    
    # 选择活动的工作表（假设是第一个表）
    sheet = workbook.active

    # 初始化一个空字典
    data_dict = {}

    # 遍历每一行，将 A 列作为键，C 列作为值，假设第一行为表头，从第二行开始
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3, values_only=True):
        key = row[0]  # A列
        value = row[1]  # B列
        data_dict[key] = value

    for filename in os.listdir(source_folder):
        if filename.endswith('.json') or filename.endswith('.Json'):
            if filename in data_dict:
                newfilename = data_dict[filename] + ".json"
                # 移动文件到新文件夹
                srcfile_path = os.path.normpath(os.path.join(source_folder, filename))
                desfile_path = os.path.normpath(os.path.join(newfolder_path, newfilename))
                desfile_path2 = os.path.join(newfolder_path, newfilename)

                try:
                    os.rename(srcfile_path, desfile_path2)
                except:
                    print(file_path+"未成功改名!")  

                pass




# 打印字典
#print(data_dict)

else:
    print("用户取消了选择文件")




