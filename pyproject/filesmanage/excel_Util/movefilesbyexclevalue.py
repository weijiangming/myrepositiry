#根据excel里记录文档名称，从文件夹中取出文件到新文件夹中
#filetitle = "_按需命名"
import os
import shutil
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import sys
from pathlib import Path
import json

# 获取当前文件的父目录
parent_dir = str(Path(__file__).resolve().parent.parent)
# 将父目录添加到sys.path
sys.path.append(parent_dir)
from filesfunction import opfiles
# 创建 Tkinter 主窗口（通常隐藏）
root = tk.Tk()
root.withdraw()

filetitle = "_取出"

# 打开文件选择对话框
file_path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel文件', '*.xlsx')])
source_folder, parent_folder = opfiles.OpFiles.select_folder()
newfolder = source_folder.split("/")[-1] + filetitle
newfolder_path = os.path.normpath(os.path.join(parent_folder, newfolder))

try:
    os.makedirs(newfolder_path)
except FileExistsError:
    pass

# 列表
data_list = []
# 检查用户是否选择了文件
if file_path:
    # 加载选定的 Excel 文件
    workbook = load_workbook(filename=file_path)
    # 选择活动的工作表
    sheet = workbook.active
    # 遍历每一行，将 A 列作为键，C 列作为值，假设第一行为表头，从第二行开始
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
        fname = row[0]  # A列
        data_list.append(fname)
else:
    print("取消了选择文件")

# 遍历源文件夹中的所有文件
for filename in os.listdir(source_folder):
    if filename.endswith('.json') or filename.endswith('.Json'):
        file_path = os.path.join(source_folder, filename)
        if filename in data_list:
            pass
            shutil.move(file_path,newfolder_path)
            

            
        


# def get_folder2fullpaths(root_dir):
#     for uziproot, uzipdirs, _ in os.walk(unzipfilepath): #xxx_解压里的文件夹
#         if not uziproot==unzipfilepath:
#             continue
#         if not os.path.exists(despath):
#             shutil.move(uzipfull_path,root2)
        
#     return folder2fullpaths