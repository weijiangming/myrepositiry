#工标库json修改重复切片
import re
import os
import sys
from pathlib import Path
import json
import openpyxl
import random

parent_dir = str(Path(__file__).resolve().parent.parent)# 获取当前文件的父目录
sys.path.append(parent_dir)# 将父目录添加到sys.path
from filesfunction import opfiles

# 搜索出“切片不带格式”和“切片带格式”中除了行首的条文编号不同，其余部分的文字内容一样的分片，进行如下处理：

# ①记录所有重复项的条文编号字段，留下第一个，其余的重复项分片删除

# ②对于去重后的分片处理：删除“切片不带格式”和“切片带格式”行首的条文编号，然后将所有重复分片各自的条文编号拿出来填被留下分片的条文编号字段中。并用“、”隔开

# ③如果被留下分片“切片不带格式”和“切片带格式”的行首没有“x.x.x”或“x.x”样式的编号，那么就将重新填的条文编号放在行首，如果这串条文编号是连续的（例如“1.0.1、1.0.2、1.0.3、1.0.4”）那么久简写成“1.0.1~1.0.4”样式；如果有不连续的号，就单独拿出来放在最后，用“、”隔开（例如“1.0.1、1.0.2、1.0.3、1.0.4、1.0.8”简写成“1.0.1~1.0.4、1.0.8”；如果被留下分片“切片不带格式”和“切片带格式”的行首有“x.x.x”或“x.x”样式的编号，那么就不用重新填编号

# ④编号和正文之间还是要添加一个空格

def simplify_versions(versions):
    # 将输入的字符串按照逗号分隔，生成版本列表，并去除多余的空格
    version_list = [v.strip() for v in versions.split('、')]
    
    # 初始化简化版本列表
    simplified = []
    
    # 初始化一个列表，用于存储连续的版本号
    current_range = []
    
    # 遍历版本列表
    for i in range(len(version_list)):
        if not current_range:  # 如果current_range为空，添加当前版本
            current_range.append(version_list[i])
        else:
            # 检查当前版本与上一个版本是否为连续版本
            current_version_parts = version_list[i].split('.')
            prev_version_parts = current_range[-1].split('.')
            
            # 比较最后一部分是否相差1，且前面部分是否相同
            if (len(current_version_parts) == len(prev_version_parts) and
                current_version_parts[:-1] == prev_version_parts[:-1] and
                int(current_version_parts[-1]) == int(prev_version_parts[-1]) + 1):
                current_range.append(version_list[i])  # 如果连续，加入current_range
            else:
                # 如果不连续，结束当前range，并存储到simplified
                if len(current_range) > 2:  # 如果连续超过两个，才简写为范围
                    simplified.append(f"{current_range[0]}~{current_range[-1]}")
                else:
                    # 如果是两个连续版本或单独版本，用“、”隔开
                    simplified.extend(current_range)
                current_range = [version_list[i]]  # 开始一个新的range
    
    # 最后一次的range处理
    if current_range:
        if len(current_range) > 2:  # 如果最后一段有超过两个版本连续
            simplified.append(f"{current_range[0]}~{current_range[-1]}")
        else:
            simplified.extend(current_range)  # 两个连续版本或单独版本
    
    return '、'.join(simplified)





# 定义文件夹路径
source_folder, parent_folder = opfiles.OpFiles.select_folder()

# 创建一个新的Excel工作簿和工作表
workbook = openpyxl.Workbook()
sheet = workbook.active

row = 0
icount = 0
icount2 = 0#测试用
icount3 = 0
icount4 = 0
icount5 = 0
icount6 = 0
# 遍历源文件夹中的所有文件
bmod = True

for jsonname in os.listdir(source_folder):
    bmod = False
    icount = icount + 1
    #定义字典 一组重复项：第一个的条文编号+其他第一个之外的其他条文编号。字典一条对应一组重复项
    articleceodefirst = {}

    #定义数组
    jsonnames = []#。json文件的文件名
    filenames = []#"文档名称"对应的值
    sliceuuids = []#"切片id"
    slicetexts = []#"切片不带格式"
    slicetext_format_list = []#"切片带格式"
    articlecodes = []#"条文编号"

    #重复的
    #reparticlecode = "" #有重复的第一条文编号
    reparticlecodes = [] #有重复的除第一条文编号之外的
    repIndex = -1

    if jsonname.endswith('.json') or jsonname.endswith('.Json'):
        file_path = os.path.join(source_folder, jsonname)
        #jsonnames.append(jsonname)
        # 打开并读取JSON文件

        # 新建一个列表，用于存储需要保留的条目
        new_data = []
        with open(file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
            icount2 = icount2 + 1
            if icount2 % 10 == 0:
                print(icount2)
        try:
            for entry in data:  
                if "文档名称" in entry and "条文编号" in entry and "切片不带格式" in entry and "切片带格式" in entry:
                    filename = entry["文档名称"]
                    slicetext = entry["切片不带格式"]
                    slicetext_format = entry["切片带格式"]
                    articlecode = entry["条文编号"]

                    newsubstring = slicetext[len(articlecode):]
                    newsubstringf = slicetext_format[len(articlecode):]
                    if len(slicetexts) > 0:
                        if newsubstring in slicetexts or newsubstringf in slicetext_format_list:
                            #说明第二次出现，重复了
                            indexT = slicetexts.index(newsubstring)
                            #获取第一次出现时的序号
                            if repIndex == -1:
                                repIndex = indexT
                            #
                            if indexT == repIndex:
                                reparticlecodes.append(articlecode)
                            
                        articlecodes.append(articlecode)
                        slicetexts.append(newsubstring)
                        slicetext_format_list.append(newsubstringf)
                    else:
                        articlecodes.append(articlecode)
                        slicetexts.append(newsubstring)
                        slicetext_format_list.append(newsubstringf) 
                else:
                    print(f'{jsonname} 字段不全缺；条文编号：{str(articlecode)}')#
            #endfor

            #开始处理一个文件
            icount3 = icount3 + 1
            # ①记录所有重复项的条文编号字段，留下第一个，其余的重复项分片删除
            for entry in data:
                # 如果条文编号不匹配，则将该条目添加到新列表
                if not entry["条文编号"] in reparticlecodes:
                    if entry["条文编号"] == articlecodes[repIndex]:
                        icount4 = icount4 + 1
                        #②对于去重后的分片处理：删除“切片不带格式”和“切片带格式”行首的条文编号，然后将所有重复分片各自的条文编号拿出来填被留下分片的条文编号字段中。并用“、”隔开
                        articlecoderep = articlecodes[repIndex]
                        newarticlecodes = articlecoderep
                        for codeT in reparticlecodes:
                            newarticlecodes = newarticlecodes + "、" + codeT
                        articlecode_T = entry["条文编号"]
                        entry["条文编号"]  = newarticlecodes

                        #修改切片

                        # ③如果被留下分片“切片不带格式”和“切片带格式”的行首没有“x.x.x”或“x.x”样式的编号，那么就将重新填的条文编号放在行首
                        # ，如果这串条文编号是连续的（例如“1.0.1、1.0.2、1.0.3、1.0.4”）那么久简写成“1.0.1~1.0.4”样式；如果有不连续的号
                        # ，就单独拿出来放在最后，用“、”隔开（例如“1.0.1、1.0.2、1.0.3、1.0.4、1.0.8”简写成“1.0.1~1.0.4、1.0.8”
                        # ；如果被留下分片“切片不带格式”和“切片带格式”的行首有“x.x.x”或“x.x”样式的编号，那么就不用重新填编号

                        # slicetexts[repIndex]
                        # slicetext_format_list[repIndex]
                        slicetext = entry["切片不带格式"]
                        slicetext_format = entry["切片带格式"]
                        articlecode = entry["条文编号"]

                        newsubstring = slicetext[len(articlecode_T):]
                        newsubstringf = slicetext_format[len(articlecode_T):]

                        newsubstring = slicetext.lstrip()
                        newsubstringf = slicetext_format.lstrip()
                        #切片剪掉"条文编号"相同的部分
                        indexfindT = -1
                        indexfindT = newsubstring.find(str(articlecode_T))
                        if indexfindT == 0:
                            newsubstring = newsubstring[len(articlecode_T):]
                        else:
                            pass

                        indexfindT = -1
                        indexfindT = newsubstringf.find(str(articlecode_T))
                        if indexfindT == 0:
                            newsubstringf = newsubstringf[len(articlecode_T):]
                        else:
                            pass


                        newsubstring2 = newsubstring.lstrip()
                        newsubstring3 = newsubstring2[:10]
                        newsubstring4 = slicetext[:20]
                        newsubstringf4 = slicetext_format[:20]
                        
                        #判断是否匹配“x.x.x”或“x.x”样式
                        bMatch = True
                        pattern = r'[A-Za-z0-9]+\.[0-9]+\.[A-Za-z0-9]+'
                        matches1 = re.findall(pattern, newsubstring3)
                        slicetextres = ""
                        slicetext_formatres = ""
                        if matches1:
                            match_res = matches1[0]
                            indexfind = -1
                            indexfind = newsubstring3.find(match_res)
                            if indexfind == 0:
                                slicetextres = newsubstring
                                slicetext_formatres = newsubstringf
                            else:
                                bMatch = False
                        else:#没有“x.x.x”或“x.x”样式的编号
                            bMatch = False

                        if not bMatch:
                            #条文编号放在行首,格式要求 像“1.0.1~1.0.4”、“1.0.1~1.0.4、1.0.8”
                            # versions = "1.0.1、1.0.2、1.0.3、1.0.4、1.0.8"
                            # result = simplify_versions(versions)
                            # print(result)  # 输出: 1.0.1~1.0.4、1.0.8
                            # versions = "b.0.1、b.0.2、b.0.3、b.0.4、b.0.8"
                            # result = simplify_versions(versions)
                            # print(result)  # 输出: b.0.1~b.0.4、b.0.8

                            try:
                                result2 = simplify_versions(newarticlecodes)
                                icount5 = icount5 + 1
                            except Exception as e:
                                print(f"An error occurred: {e}")
                                print(newarticlecodes)
                                print(jsonname)
                                icount6 = icount6 + 1
                            
                            #保证后面有空格
                            
                            if newsubstring and newsubstring[0] == ' ':
                                slicetextres = result2 + newsubstring
                            else:
                                slicetextres = result2 + ' ' + newsubstring
                                
                            #test
                            if  len(newsubstring) == 0:
                                pass
                            
                            if newsubstringf and newsubstringf[0] == ' ':
                                slicetext_formatres = result2 + newsubstringf
                            else:
                                slicetext_formatres = result2 + ' '+ newsubstringf

                             #记录excel
                            record1 = slicetextres[:50]
                            record2 = slicetext_formatres[:50]
                            row += 1
                            sheet.cell(row=row, column=1, value=jsonname)
                            sheet.cell(row=row, column=2, value=newarticlecodes)
                            sheet.cell(row=row, column=3, value=result2)
                            sheet.cell(row=row, column=4, value=newsubstring4)
                            sheet.cell(row=row, column=5, value=record1)
                            sheet.cell(row=row, column=6, value=newsubstringf4)
                            sheet.cell(row=row, column=7, value=record2)

                        else:
                            #有匹配到有“x.x.x”或“x.x”样式的编号
                            row += 1
                            sheet.cell(row=row, column=9, value=jsonname)
                            sheet.cell(row=row, column=10, value=newsubstring4)
                            sheet.cell(row=row, column=11, value=newsubstringf4)
                            sheet.cell(row=row, column=12, value="有匹配到，不需修改")

                        entry["切片不带格式"] = slicetextres
                        entry["切片带格式"] = slicetext_formatres
                    new_data.append(entry)

            # 将修改后的数据写回文件
            with open(file_path, 'w', encoding='utf-8') as json_file:
                json.dump(new_data, json_file, ensure_ascii=False, indent=4)


        except json.JSONDecodeError:
            print(f'{jsonname} 该文件单独查是什么问题')
pass


excel_file_name = source_folder.split("/")[-1] + "_条文说明略写前后对比记录.xlsx"
excelfolder_path = os.path.normpath(os.path.join(parent_folder, excel_file_name))
workbook.save(excelfolder_path)

