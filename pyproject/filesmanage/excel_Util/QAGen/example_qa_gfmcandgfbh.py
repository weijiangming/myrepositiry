##测试用QA用例：规范名称+规范编号
import re
import os
import sys
from pathlib import Path
import json
import openpyxl

# 获取当前文件的父目录
parent_dir = str(Path(__file__).resolve().parent.parent)
# 将父目录添加到sys.path
sys.path.append(parent_dir)

from filesfunction import opfiles

# 定义文件夹路径
source_folder, parent_folder = opfiles.OpFiles.select_folder()

filenames = []
jsonnames = []
articlecodes = []
articles = []
icount = 0
icount2 = 0
# 遍历源文件夹中的所有文件
for filename in os.listdir(source_folder):
    icount = icount + 1  
    if filename.endswith('.json') or filename.endswith('.Json'):
        file_path = os.path.join(source_folder, filename)
        #jsonnames.append(filename)
        # 打开并读取JSON文件
        with open(file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
  
        try:
            for entry in data:   
                if "文档名称" in entry and "条文编号" in entry and "切片不带格式" in entry:
                    name = entry["文档名称"]
                    article = entry["切片不带格式"]
                    articlecode = entry["条文编号"]
                    if "1.0.1" in str(articlecode):
                        icount2 = icount2 + 1 
                        filenames.append(name)
                        jsonnames.append(filename)
                        articlecodes.append("1.0.1")
                        articles.append(article)
                        break
                else:
                    print("entry does not have a '版本' key")
            

        except json.JSONDecodeError:
            print(f'{filename} 不是有效的 JSON 文件')
    if not icount2 == icount:
        pass

# 创建一个新的Excel工作簿和工作表
workbook = openpyxl.Workbook()
sheet = workbook.active

#规范名称 1 如工：程结构通用规范
row = 0
for item in filenames:
    result = ""
    result2 = ""
    row += 1

    pattern = r"》(.*)"
    match = re.search(pattern, item)

    if match:
        result = match.group(1)
        print(result)  # 输出：CJJ 83-2016
    else:
        result = "XXXXXXXX"


    pattern = r"《(.*?)》"
    match2 = re.search(pattern, item)
    if match2:
        result2 = match2.group(1)

    result3 = result2 + " " + result
    sheet.cell(row=row, column=1, value=result3)
    
#答案 2  如：1.01
row = 0

for item in articlecodes:
    row += 1
    sheet.cell(row=row, column=2, value=item)

#截图或内容分片 3
row = 0
for item in articles:
    row += 1
    sheet.cell(row=row, column=3, value=item)

#规范名称及编号 4 如：《工程结构通用规范》GB55001-2021
row = 0
for item in filenames:
    row += 1
    sheet.cell(row=row, column=4, value=item)

excel_file_name = source_folder.split("/")[-1] + "_提取文件名.xlsx"
excelfolder_path = os.path.normpath(os.path.join(parent_folder, excel_file_name))
workbook.save(excelfolder_path)