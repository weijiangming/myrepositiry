#工标库json条文编号有漏的拎出来并记录到文件夹
import re
import os
import sys
from pathlib import Path
import json
import openpyxl
import random
import inspect

parent_dir = str(Path(__file__).resolve().parent.parent.parent)# 获取当前文件的父目录
sys.path.append(parent_dir)# 将父目录添加到sys.path
parent_dir = str(Path(__file__).resolve().parent.parent)# 获取当前文件的父目录
sys.path.append(parent_dir)# 将父目录添加到sys.path
from filesfunction import opfiles
from filenamehelpers import filenamesort

def get_string_before_last_dot(input_string):
    # 找到最后一个"."的位置
    last_dot_index = input_string.rfind('.')
    # 如果找到了"."，返回"."之前的部分；如果没有找到，返回原始字符串
    if last_dot_index != -1:
        return input_string[:last_dot_index]
    else:
        return input_string

def find_missing_versions(versions):
    split_versions = [re.split(r'([A-Z]?)$', version) for version in versions]
    groups = {}
    
    for version_parts in split_versions:
        base = version_parts[0]
        letter = version_parts[1]
        if base not in groups:
            groups[base] = []
        groups[base].append(letter)
    
    missing_versions = []
    
    for base, letters in groups.items():
        if letters[0] == '':
            # Find missing numeric versions
            last_part_numbers = sorted(int(base.split('.')[-1]) for base in versions)
            for i in range(last_part_numbers[0], last_part_numbers[-1]):
                if i not in last_part_numbers:
                    #missing_versions.append(f"{base[:-1]}{i}")
                    missing_versions.append(f"{'.'.join(base.split('.')[:-1])}.{i}")
        else:
            # Find missing alphabetic versions
            letters = sorted(letters)
            for i in range(ord(letters[0]), ord(letters[-1])):
                if chr(i) not in letters:
                    missing_versions.append(f"{base}{chr(i)}")

    return list(set(missing_versions))  # Avoid redundant outputs

def find_missing_from_string(version_string):
    # Split the string into individual versions by the separator '、'
    versions = version_string.split('、')

    # Process the list of versions to find missing versions
    return find_missing_versions(versions)

# 定义文件夹路径
source_folder, parent_folder = opfiles.OpFiles.select_folder()

# 创建一个新的Excel工作簿和工作表
workbook = openpyxl.Workbook()
sheet = workbook.active

linkkey = '_' 
row = 0
icount = 0

icount3 = 0
icount4 = 0
icount5 = 0
icount6 = 0
# 遍历源文件夹中的所有文件
bmod = True

for jsonname in os.listdir(source_folder):

    bmod = False
    #定义数组
    jsonnames = []#。json文件的文件名
    filenames = []#"文档名称"对应的值
    sliceuuids = []#"切片id"
    slicetexts = []#"切片不带格式"
    slicetexts_f = []#"切片带格式"
    
    articlecodes = []#"条文编号"

    #定义
    #定义字典 一个重复组：重复分片的第一项的序号+对应的所有重复条文编号最后的"."之前的部分
    index2front_dict = {}
    #重复组条文编号最后的"."之前的部分+重复组第一项以外的其他项的条文编号的数组（frontdot2codelist）；字典一项对应一个重复组；字典由多个重复组组成。
    frontindex_codes_dictdel = {}
    if jsonname.endswith('.json') or jsonname.endswith('.Json'):
        file_path = os.path.join(source_folder, jsonname)
        #jsonnames.append(jsonname)
        # 打开并读取JSON文件

        # 新建一个列表，用于存储需要保留的条目
        sync_index = -1
        new_data = []
        with open(file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
            icount = icount + 1
            if icount % 100 == 0:
                print(icount)
        
        #二级标题分组
        articlecodesgroup = []#"条文编号"
        frontcodepre = ""#上一个二级标题
        frontindex_codes_dict = {}

        #条文编号上级标题
        fcode2codes_dict = {}
        # articlecode
        content2codes_dict = {} #"切片不带格式"和"条文编号"列表的字典
        code2index_dict = {}    #"条文编号"和sync_index

        try:
            for entry in data:  
                if "文档名称" in entry and "条文编号" in entry and "切片不带格式" in entry and "切片带格式" in entry:
                    sync_index = sync_index + 1

                    filename = entry["文档名称"]
                    slicetext = entry["切片不带格式"]
                    slicetext_format = entry["切片带格式"]
                    articlecode = entry["条文编号"]
                    frontcode = get_string_before_last_dot(articlecode)

                    if frontcode not in fcode2codes_dict:
                        fcode2codes_dict[frontcode] = []
                    if articlecode not in fcode2codes_dict[frontcode]:
                        fcode2codes_dict[frontcode].append(articlecode)
                else:
                    print(f'{jsonname} 字段不全缺；条文编号：{str(articlecode)}')#

            #开始处理一个文件
            icount3 = icount3 + 1
            # ①记录所有重复项的条文编号字段，留下第一个，其余的重复项分片删除
            
            sync_index = -1#确保第一个的序号是零
            for fcode, codes in fcode2codes_dict.items():
                codesless = find_missing_versions(codes)

                #确保和上一次for entry in data:规则相同，保证sync_index序号相同。
                icount6 = icount6 + 1
                break
                if "文档名称" in entry and "条文编号" in entry and "切片不带格式" in entry and "切片带格式" in entry:
                    sync_index = sync_index + 1
                
                twbhT = str(entry["条文编号"])
                frontcode = get_string_before_last_dot(twbhT)

                #获取twbhT在content2codes_dict里的条文编号数组
                code_list = []
                for key, value in content2codes_dict.items():
                    if twbhT in value:
                        code_list.extend(value)
                        break 
                #test
                if twbhT == "8.0.1":
                    pass

                issave = True
                if len(code_list) == 1:
                    issave = True
                elif len(code_list) > 1:
                    frontcode2codes_dict = {}
                    for code in code_list:
                        frontcodeT = get_string_before_last_dot(code)
                        frontcode2codes_dict
                        if frontcodeT not in frontcode2codes_dict:
                            frontcode2codes_dict[frontcodeT] = []
                        frontcode2codes_dict[frontcodeT].append(code)

                    #按frontcode2codes_dict开始处理
                    newarticlecodes = ""
                    for frontcode, subcodelist in frontcode2codes_dict.items():
                        index = -1
                        try:
                            index = subcodelist.index(twbhT)
                            
                            if index == 0:
                                for codett in subcodelist:
                                    if newarticlecodes == "":
                                        newarticlecodes = codett
                                    else:
                                        newarticlecodes = newarticlecodes + "、" + codett
                                entry["条文编号"]  = newarticlecodes

                                slicetext = entry["切片不带格式"]
                                slicetext_format = entry["切片带格式"]
                                #articlecode = entry["条文编号"]

                                newsubstring = slicetext.lstrip()
                                newsubstringf = slicetext_format.lstrip()
                                #
                                indexfindT = -1
                                indexfindT = newsubstring.find(twbhT)
                                if indexfindT == 0:
                                    newsubstring = newsubstring[len(twbhT):]
                                else:
                                    pass

                                indexfindT = -1
                                indexfindT = newsubstringf.find(twbhT)
                                if indexfindT == 0:
                                    newsubstringf = newsubstringf[len(twbhT):]
                                else:
                                    pass
                                #取前10的字符用于判断是否有“x.x.x”或“x.x”样式                           
                                strforpattern = newsubstring[:10]
                                strforpattern = strforpattern.lstrip()

                                #判断是否匹配“x.x.x”或“x.x”样式
                                bMatch = True
                                pattern = r'[A-Za-z0-9]+\.[0-9]+\.[A-Za-z0-9]+'
                                matches1 = re.findall(pattern, strforpattern)
                                slicetextres = ""
                                slicetext_formatres = ""
                                if matches1:
                                    match_res = matches1[0]
                                    indexfind = -1
                                    indexfind = strforpattern.find(match_res)
                                    if indexfind == 0:
                                        slicetextres = newsubstring
                                        slicetext_formatres = newsubstringf
                                    else:
                                        bMatch = False
                                else:#没有“x.x.x”或“x.x”样式的编号
                                    if bMatch:
                                        pattern2 = r'[A-Za-z0-9]+\.[A-Za-z0-9]+'
                                        matches2 = re.findall(pattern2, strforpattern)
                                        if matches2:
                                            match_res = matches2[0]
                                            indexfind = -1
                                            indexfind = strforpattern.find(match_res)
                                            if indexfind == 0:
                                                slicetextres = newsubstring
                                                slicetext_formatres = newsubstringf
                                            else:
                                                bMatch = False
                                        else:
                                            bMatch = False

                                #用于记录excel非功能业务
                                newsubstring4 = slicetext[:20]
                                newsubstringf4 = slicetext_format[:20]

                                if not bMatch:
                                    #条文编号放在行首,格式要求 像“1.0.1~1.0.4”、“1.0.1~1.0.4、1.0.8”
                                    # versions = "1.0.1、1.0.2、1.0.3、1.0.4、1.0.8"
                                    # result = simplify_versions(versions)
                                    # print(result)  # 输出: 1.0.1~1.0.4、1.0.8
                                    # versions = "b.0.1、b.0.2、b.0.3、b.0.4、b.0.8"
                                    # result = simplify_versions(versions)
                                    # print(result)  # 输出: b.0.1~b.0.4、b.0.8
                                    try:
                                        result2 = filenamesort.OpFileName.simplify_versions(newarticlecodes)
                                        icount5 = icount5 + 1
                                    except Exception as e:
                                        print(f"An error occurred: {e}")
                                        print(newarticlecodes)
                                        print(jsonname)
                                       # icount6 = icount6 + 1
                                    
                                    #保证后面有空格
                                    if newsubstring and newsubstring[0] == ' ':
                                        slicetextres = result2 + newsubstring
                                    else:
                                        slicetextres = result2 + ' ' + newsubstring
                                        
                                    #test
                                    if  len(newsubstring) == 0:
                                        pass
                                    
                                    if newsubstringf and newsubstringf[0] == ' ':
                                        slicetext_formatres = result2 + newsubstringf
                                    else:
                                        slicetext_formatres = result2 + ' '+ newsubstringf


                                    record1 = slicetextres[:50]
                                    record2 =  slicetext_formatres[:50]
                                    row += 1
                                    sheet.cell(row=row, column=1, value=jsonname)
                                    sheet.cell(row=row, column=2, value=newarticlecodes)
                                    sheet.cell(row=row, column=3, value=result2)
                                    sheet.cell(row=row, column=4, value=newsubstring4)
                                    sheet.cell(row=row, column=5, value=record1)
                                    sheet.cell(row=row, column=6, value=newsubstringf4)
                                    sheet.cell(row=row, column=7, value=record2)

                                else:
                                    #有匹配到有“x.x.x”或“x.x”样式的编号
                                    row += 1
                                    sheet.cell(row=row, column=9, value=jsonname)
                                    sheet.cell(row=row, column=10, value=newsubstring4)
                                    sheet.cell(row=row, column=11, value=newsubstringf4)
                                    sheet.cell(row=row, column=12, value="有匹配到，不需修改")

                                entry["切片不带格式"] = slicetextres
                                entry["切片带格式"] = slicetext_formatres
                                        
                                break
                            elif index > 0:
                                issave = False
                                break           
                        except ValueError:
                            pass
                if issave:
                    new_data.append(entry)

            # 将修改后的数据写回文件
            with open(file_path, 'w', encoding='utf-8') as json_file:
                json.dump(new_data, json_file, ensure_ascii=False, indent=4)

        except json.JSONDecodeError:
            print(f'{jsonname} 该文件单独查是什么问题')
pass
print("icount6")
print(icount6)
 
# excel_file_name = source_folder.split("/")[-1] + "_条文说明略写前后对比记录.xlsx"
# excelfolder_path = os.path.normpath(os.path.join(parent_folder, excel_file_name))
# workbook.save(excelfolder_path)

