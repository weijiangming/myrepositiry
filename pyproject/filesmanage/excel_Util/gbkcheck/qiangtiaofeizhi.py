#加强条废止字段
import os
import sys
from pathlib import Path
import json
import openpyxl
from openpyxl import load_workbook
import pandas as pd

parent_dir = str(Path(__file__).resolve().parent.parent.parent)# 获取当前文件的父目录
sys.path.append(parent_dir)# 将父目录添加到sys.path
parent_dir = str(Path(__file__).resolve().parent.parent)# 获取当前文件的父目录
sys.path.append(parent_dir)# 将父目录添加到sys.path
from filesfunction import opfiles

# 定义文件夹路径
source_folder, parent_folder = opfiles.OpFiles.select_folder()

# 创建一个新的Excel工作簿和工作表
workbook = openpyxl.Workbook()
sheet = workbook.active

row = 0
icount = 0
icount2 = 0 #测试用
icount3 = 0
icount4 = 0 
icount5 = 0
icount6 = 0
# 遍历源文件夹中的所有文件
bmod = True

for jsonname in os.listdir(source_folder):
    if jsonname.endswith('.xlsx') :
        file_path = os.path.join(source_folder, jsonname)
        df = pd.read_excel(file_path, skiprows=1, header=None, usecols="A:E")
        print(df)

        for row in df.itertuples(index=True, name='Row'):
            # row 是一个 namedtuple 对象，你可以通过列名访问每个值
            # 注意：列名需要以 '_Column1' 的形式访问
            print(f"Row {row.Index}: {row._1}, {row._2}, {row._3}, {row._4}, {row._5}")

        pass

        # workbook = load_workbook(filename=file_path)
    
        # # 选择活动的工作表（假设是第一个表）
        # sheet = workbook.active
        # # 遍历每一行，将 A 列作为键，C 列作为值，假设第一行为表头，从第二行开始
        # for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3, values_only=True):
        #     key = row[0]  # A列
        #     value = row[1]  # B列

exit()

for jsonname in os.listdir(source_folder):
    bmod = False

    #定义数组
    jsonnames = []#。json文件的文件名
    filenames = []#"文档名称"对应的值
    sliceuuids = []#"切片id"
    slicetexts = []#"切片不带格式"
    slicetext_format_list = []#"切片带格式"
    articlecodes = []#"条文编号"

    #重复的
     #定义字典 一个重复组：重复组第一项的序号+重复组第一项以外i的其他项的条文编号的数组（Index2codelist）；字典一项对应一个重复组；字典由多个重复组组成。
    repeat_dict = {}
    if jsonname.endswith('.json') or jsonname.endswith('.Json'):
        file_path = os.path.join(source_folder, jsonname)
        # 打开并读取JSON文件
        new_data = []
        with open(file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
            icount = icount + 1
            if icount % 100 == 0:
                print(icount)
        try:
            for entry in data:  
                if "文档名称" in entry and "条文编号" in entry and "切片不带格式" in entry and "切片带格式" in entry:
                    icount2 = icount2 + 1
                    filename = entry["文档名称"]
                    slicetext = entry["切片不带格式"]
                    slicetext_format = entry["切片带格式"]
                    articlecode = entry["条文编号"]
                    sliceuuid = entry["切片id"]

                    is_strong = 0
                    aid_stronglist = []
                    is_repeal = 1
                    aid_repeallist = []

                    aid_repeal = articlecode + " " + "自2025年1月1日起废止该条，根据《施工现场临时用电安全技术规范》JGJ 46-2005 标识废止"
                    aid_repeallist.append(aid_repeal) 
                                       
                    entry["is_strong"] = is_strong
                    entry["aid_strong"] = aid_stronglist
                    entry["is_repeal"] = is_repeal
                    entry["aid_repeal"] = aid_repeallist

                else:
                    print(f'{jsonname} 字段不全缺；条文编号：{str(articlecode)}')

            
        except json.JSONDecodeError:
            print(f'{jsonname} 该文件单独查是什么问题')

        with open(file_path, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=4)


   

