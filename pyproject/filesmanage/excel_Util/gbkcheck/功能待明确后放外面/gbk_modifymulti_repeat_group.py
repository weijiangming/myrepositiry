#工标库json修改重复切片；执行该文件之前先去条纹编号和切片都相同的情况
import re
import os
import sys
from pathlib import Path
import json
import openpyxl
import random
import inspect

parent_dir = str(Path(__file__).resolve().parent.parent)# 获取当前文件的父目录
sys.path.append(parent_dir)# 将父目录添加到sys.path
from filesfunction import opfiles

# 搜索出“切片不带格式”和“切片带格式”中除了行首的条文编号不同，其余部分的文字内容一样的分片，进行如下处理：

# ①记录所有重复项的条文编号字段，留下第一个，其余的重复项分片删除

# ②对于去重后的分片处理：删除“切片不带格式”和“切片带格式”行首的条文编号，然后将所有重复分片各自的条文编号拿出来填被留下分片的条文编号字段中。并用“、”隔开

# ③如果被留下分片“切片不带格式”和“切片带格式”的行首没有“x.x.x”或“x.x”样式的编号，那么就将重新填的条文编号放在行首，如果这串条文编号是连续的（例如“1.0.1、1.0.2、1.0.3、1.0.4”）那么久简写成“1.0.1~1.0.4”样式；如果有不连续的号，就单独拿出来放在最后，用“、”隔开（例如“1.0.1、1.0.2、1.0.3、1.0.4、1.0.8”简写成“1.0.1~1.0.4、1.0.8”；如果被留下分片“切片不带格式”和“切片带格式”的行首有“x.x.x”或“x.x”样式的编号，那么就不用重新填编号

# ④编号和正文之间还是要添加一个空格


def simplify_versions(versions):
    # 将输入的字符串按照逗号分隔，生成版本列表，并去除多余的空格
    version_list = [v.strip() for v in versions.split('、')]
    
    # 初始化简化版本列表
    simplified = []
    
    # 初始化一个列表，用于存储连续的版本号
    current_range = []
    
    # 遍历版本列表
    for i in range(len(version_list)):
        if not current_range:  # 如果current_range为空，添加当前版本
            current_range.append(version_list[i])
        else:
            # 检查当前版本与上一个版本是否为连续版本
            current_version_parts = version_list[i].split('.')
            prev_version_parts = current_range[-1].split('.')
            
            # 比较最后一部分是否相差1，且前面部分是否相同
            if (len(current_version_parts) == len(prev_version_parts) and
                current_version_parts[:-1] == prev_version_parts[:-1] and
                int(current_version_parts[-1]) == int(prev_version_parts[-1]) + 1):
                current_range.append(version_list[i])  # 如果连续，加入current_range
            else:
                # 如果不连续，结束当前range，并存储到simplified
                if len(current_range) > 2:  # 如果连续超过两个，才简写为范围
                    simplified.append(f"{current_range[0]}~{current_range[-1]}")
                else:
                    # 如果是两个连续版本或单独版本，用“、”隔开
                    simplified.extend(current_range)
                current_range = [version_list[i]]  # 开始一个新的range
    
    # 最后一次的range处理
    if current_range:
        if len(current_range) > 2:  # 如果最后一段有超过两个版本连续
            simplified.append(f"{current_range[0]}~{current_range[-1]}")
        else:
            simplified.extend(current_range)  # 两个连续版本或单独版本
    
    return '、'.join(simplified)

def get_string_before_last_dot(input_string):
    # 找到最后一个"."的位置
    last_dot_index = input_string.rfind('.')
    # 如果找到了"."，返回"."之前的部分；如果没有找到，返回原始字符串
    if last_dot_index != -1:
        return input_string[:last_dot_index]
    else:
        return input_string

#index是否在indexfront_codes_dict
def isIndexInDictKey(index, dict, linksymbol):
    frontlist = []
    found = False
    for key in dict:
        try:
            indexT = key.split(linksymbol)[0]
            if index == int(indexT):
                found = True
                frontlist.append(key.split(linksymbol)[1]) 

        except IndexError:
            continue
    return found, frontlist
    #使用方法：found, frontlist = 

#frontdotcode是否在indexfront_codes_dict
def isfrontcodeInDictValue(frontcode, dict, linksymbol):
    
    found = False
    indexlist = []
    for key in dict:
        try:
            frontcodeT = key.split(linksymbol)[1]
            if frontcode == frontcodeT:
                found = True
                indexlist.append(key.split(linksymbol)[0])
                
        except IndexError:
            continue
    return found, indexlist

# 定义文件夹路径
source_folder, parent_folder = opfiles.OpFiles.select_folder()

# 创建一个新的Excel工作簿和工作表
workbook = openpyxl.Workbook()
sheet = workbook.active

linkkey = '_' 
row = 0
icount = 0

icount3 = 0
icount4 = 0
icount5 = 0
icount6 = 0
# 遍历源文件夹中的所有文件
bmod = True

for jsonname in os.listdir(source_folder):

    bmod = False
    #定义数组
    jsonnames = []#。json文件的文件名
    filenames = []#"文档名称"对应的值
    sliceuuids = []#"切片id"
    slicetexts = []#"切片不带格式"
    slicetext_format_list = []#"切片带格式"
    articlecodes = []#"条文编号"

    #定义重复的
    #定义字典 一个重复组：重复分片的第一项的序号+对应的所有重复条文编号最后的"."之前的部分
    index2front_dict = {}
    #重复组条文编号最后的"."之前的部分+重复组第一项以外的其他项的条文编号的数组（frontdot2codelist）；字典一项对应一个重复组；字典由多个重复组组成。
    indexfront_codes_dict = {}
    if jsonname.endswith('.json') or jsonname.endswith('.Json'):
        file_path = os.path.join(source_folder, jsonname)
        #jsonnames.append(jsonname)
        # 打开并读取JSON文件

        # 新建一个列表，用于存储需要保留的条目
        sync_index = -1
        new_data = []
        with open(file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
            icount = icount + 1
            if icount % 10 == 0:
                print(icount)
        
        #二级标题分组
        articlecodesgroup = []#"条文编号"
        try:
            for entry in data:  
                if "文档名称" in entry and "条文编号" in entry and "切片不带格式" in entry and "切片带格式" in entry:
                    sync_index = sync_index + 1
                    filename = entry["文档名称"]
                    slicetext = entry["切片不带格式"]
                    slicetext_format = entry["切片带格式"]
                    articlecode = entry["条文编号"]

                    newsubstring = slicetext
                    newsubstringf = slicetext_format
                     #切片剪掉"条文编号"相同的部分
                    indexfindT = -1
                    indexfindT = newsubstring.find(str(articlecode))
                    if indexfindT == 0:
                        newsubstring = newsubstring[len(articlecode):]
                    else:
                        pass

                    indexfindT = -1
                    indexfindT = newsubstringf.find(str(articlecode))
                    if indexfindT == 0:
                        newsubstringf = newsubstringf[len(articlecode):]
                    else:
                        pass

                    #test
                    twbhT = entry["条文编号"]
                    if twbhT == "4.2.6":#"4.5.5":
                        pass

                    if newsubstring in slicetexts:
                        #进入这里切片重复了
                        indexT = slicetexts.index(newsubstring)#有重复的切片正在数组里的首序号
                        frontdot = get_string_before_last_dot(articlecode)
                       
                        #重复4种情况：就是index 和 front和字典的4种情况：
                        #1.index和front都在；2.index在和front不在；3.index不在和front在；4.index和front都不在；
                        #情况1、2。
                        found, frontlist = isIndexInDictKey(indexT,indexfront_codes_dict,linkkey)
                        if found:#重复组已有在字典记录  
                            #情况1.index和front都在
                            if frontdot in frontlist:
                                indexfront = str(indexT) + linkkey + frontdot 
                                try:
                                    indexfront_codes_dict[indexfront].append(articlecode)
                                except Exception as e:
                                    print(f"An error occurred: {e}")
                                    print(len(indexfront_codes_dict))
                                    print(jsonname)

                            else:#情况2.index在和front不在
                                foundtwbh = articlecodes[indexT]
                                foundfront = get_string_before_last_dot(foundtwbh)
                                if foundfront != frontdot:
                                    indexfront = str(sync_index) + linkkey + frontdot
                                    indexfront_codes_dict[indexfront] = [] 
                                else:
                                    pass #这种情况待调试补上
                                    #indexfront_codes_dict[indexfront].append(articlecode)
                        else:
                            #情况3index不在和front在
                            if frontdot in frontlist:
                                indexfront = str(indexT) + linkkey + frontdot
                                indexfront_codes_dict[indexfront] = []
                                #indexfront_codes_dict[indexfront].append(articlecode)
                                index2front_dict[indexT] = frontdot
                            else:#情况4 跨组的切片相同：条件：indexT的front和当前frontcode不同，不跨组则相同
                                foundtwbh = articlecodes[indexT]
                                foundfront = get_string_before_last_dot(foundtwbh)
                                if foundfront != frontdot:#不是在当前组找到 只是当前组有于其他组相同的切片；但是字典里记录下，不保证当前组会有与sync_index相同的切片
                                    indexfront = str(sync_index) + linkkey + frontdot
                                    indexfront_codes_dict[indexfront] = []
                                    #indexfront_codes_dict[indexfront].append(articlecode)
                                else:
                                    indexfront = str(indexT) + linkkey + frontdot
                                    indexfront_codes_dict[indexfront] = []
                                    indexfront_codes_dict[indexfront].append(articlecode)

                    articlecodes.append(articlecode)
                    slicetexts.append(newsubstring)
                    slicetext_format_list.append(newsubstringf)
                    
                else:
                    print(f'{jsonname} 字段不全缺；条文编号：{str(articlecode)}')#
            #endfor

            #开始处理一个文件
            icount3 = icount3 + 1
            # ①记录所有重复项的条文编号字段，留下第一个，其余的重复项分片删除
            
            sync_index = -1#确保第一个的序号是零
            for entry in data:
                #确保和上一次for entry in data:规则相同，保证序号相同。
                if "文档名称" in entry and "条文编号" in entry and "切片不带格式" in entry and "切片带格式" in entry:
                    sync_index = sync_index + 1
                
                #test
                twbhT = entry["条文编号"]
                if twbhT == "4.3.5":
                    pass
                
                #
                found = False
                twbhT = entry["条文编号"]
                frontdot = get_string_before_last_dot(twbhT)
                for key, value_list in indexfront_codes_dict.items():
                    if twbhT in value_list:
                        found = True
                        break 

                if found == False:#需要保留的都在new_data里
                    #只处理sync_index in inidex2front_dict对应的分片，其他不用动，因为整个entry会加入new_data：new_data.append(entry)
                    indexfront = str(sync_index) + linkkey + frontdot
                    if indexfront in indexfront_codes_dict:
                        icount4 = icount4 + 1
                        #②对于去重后的分片处理：删除“切片不带格式”和“切片带格式”行首的条文编号，然后将所有重复分片各自的条文编号拿出来填被留下分片的条文编号字段中。并用“、”隔开
                        articlecoderep = articlecodes[sync_index]
                        newarticlecodes = articlecoderep
                        twbhlist = indexfront_codes_dict[indexfront]
                        for codeT in twbhlist:
                            newarticlecodes = newarticlecodes + "、" + codeT
                        articlecode_T = entry["条文编号"]
                        entry["条文编号"]  = newarticlecodes

                        #修改切片
                        # ③如果被留下分片“切片不带格式”和“切片带格式”的行首没有“x.x.x”或“x.x”样式的编号，那么就将重新填的条文编号放在行首
                        # ，如果这串条文编号是连续的（例如“1.0.1、1.0.2、1.0.3、1.0.4”）那么久简写成“1.0.1~1.0.4”样式；如果有不连续的号
                        # ，就单独拿出来放在最后，用“、”隔开（例如“1.0.1、1.0.2、1.0.3、1.0.4、1.0.8”简写成“1.0.1~1.0.4、1.0.8”
                        # ；如果被留下分片“切片不带格式”和“切片带格式”的行首有“x.x.x”或“x.x”样式的编号，那么就不用重新填编号

                        slicetext = entry["切片不带格式"]
                        slicetext_format = entry["切片带格式"]
                        #articlecode = entry["条文编号"]

                        newsubstring = slicetext.lstrip()
                        newsubstringf = slicetext_format.lstrip()
                        #
                        indexfindT = -1
                        indexfindT = newsubstring.find(str(articlecode_T))
                        if indexfindT == 0:
                            newsubstring = newsubstring[len(articlecode_T):]
                        else:
                            pass

                        indexfindT = -1
                        indexfindT = newsubstringf.find(str(articlecode_T))
                        if indexfindT == 0:
                            newsubstringf = newsubstringf[len(articlecode_T):]
                        else:
                            pass

                        newsubstring2 = newsubstring.lstrip()
                        newsubstring3 = newsubstring2[:10]
                        newsubstring4 = slicetext[:20]
                        newsubstringf4 = slicetext_format[:20]
                        
                        #判断是否匹配“x.x.x”或“x.x”样式
                        bMatch = True
                        pattern = r'[A-Za-z0-9]+\.[0-9]+\.[A-Za-z0-9]+'
                        matches1 = re.findall(pattern, newsubstring3)
                        slicetextres = ""
                        slicetext_formatres = ""
                        if matches1:
                            match_res = matches1[0]
                            indexfind = -1
                            indexfind = newsubstring3.find(match_res)
                            if indexfind == 0:
                                slicetextres = newsubstring
                                slicetext_formatres = newsubstringf
                            else:
                                bMatch = False
                        else:#没有“x.x.x”或“x.x”样式的编号
                            if bMatch:
                                pattern2 = r'[A-Za-z0-9]+\.[A-Za-z0-9]+'
                                matches2 = re.findall(pattern2, newsubstring3)
                                if matches2:
                                    match_res = matches2[0]
                                    indexfind = -1
                                    indexfind = newsubstring3.find(match_res)
                                    if indexfind == 0:
                                        slicetextres = newsubstring
                                        slicetext_formatres = newsubstringf
                                    else:
                                        bMatch = False
                                else:
                                    bMatch = False

                        if not bMatch:
                            #条文编号放在行首,格式要求 像“1.0.1~1.0.4”、“1.0.1~1.0.4、1.0.8”
                            # versions = "1.0.1、1.0.2、1.0.3、1.0.4、1.0.8"
                            # result = simplify_versions(versions)
                            # print(result)  # 输出: 1.0.1~1.0.4、1.0.8
                            # versions = "b.0.1、b.0.2、b.0.3、b.0.4、b.0.8"
                            # result = simplify_versions(versions)
                            # print(result)  # 输出: b.0.1~b.0.4、b.0.8

                            try:
                                result2 = simplify_versions(newarticlecodes)
                                icount5 = icount5 + 1
                            except Exception as e:
                                print(f"An error occurred: {e}")
                                print(newarticlecodes)
                                print(jsonname)
                                icount6 = icount6 + 1
                            
                            #保证后面有空格
                            
                            if newsubstring and newsubstring[0] == ' ':
                                slicetextres = result2 + newsubstring
                            else:
                                slicetextres = result2 + ' ' + newsubstring
                                
                            #test
                            if  len(newsubstring) == 0:
                                pass
                            
                            if newsubstringf and newsubstringf[0] == ' ':
                                slicetext_formatres = result2 + newsubstringf
                            else:
                                slicetext_formatres = result2 + ' '+ newsubstringf

                             #记录excel
                            record1 = slicetextres[:50]
                            record2 = slicetext_formatres[:50]
                            row += 1
                            sheet.cell(row=row, column=1, value=jsonname)
                            sheet.cell(row=row, column=2, value=newarticlecodes)
                            sheet.cell(row=row, column=3, value=result2)
                            sheet.cell(row=row, column=4, value=newsubstring4)
                            sheet.cell(row=row, column=5, value=record1)
                            sheet.cell(row=row, column=6, value=newsubstringf4)
                            sheet.cell(row=row, column=7, value=record2)

                        else:
                            #有匹配到有“x.x.x”或“x.x”样式的编号
                            row += 1
                            sheet.cell(row=row, column=9, value=jsonname)
                            sheet.cell(row=row, column=10, value=newsubstring4)
                            sheet.cell(row=row, column=11, value=newsubstringf4)
                            sheet.cell(row=row, column=12, value="有匹配到，不需修改")

                        entry["切片不带格式"] = slicetextres
                        entry["切片带格式"] = slicetext_formatres
                    new_data.append(entry)

            # 将修改后的数据写回文件
            with open(file_path, 'w', encoding='utf-8') as json_file:
                json.dump(new_data, json_file, ensure_ascii=False, indent=4)


        except json.JSONDecodeError:
            print(f'{jsonname} 该文件单独查是什么问题')
pass


excel_file_name = source_folder.split("/")[-1] + "_条文说明略写前后对比记录.xlsx"
excelfolder_path = os.path.normpath(os.path.join(parent_folder, excel_file_name))
workbook.save(excelfolder_path)

