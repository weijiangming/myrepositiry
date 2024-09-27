#工标库检查json有某种情况拎出来另一个文件夹
#条文说明JSON文件中 存在“切片不带格式”和“切片带格式”中4.1、4.2、4.3、4.4、4.5、46、47、4.8、47、4.9、4.10、4.11、4.12、4.13 4.1这种情况的文件挑选出来
import re
import os
import sys
from pathlib import Path
import json
import shutil
import openpyxl
import random

parent_dir = str(Path(__file__).resolve().parent.parent)# 获取当前文件的父目录
sys.path.append(parent_dir)# 将父目录添加到sys.path
from filesfunction import opfiles



# 定义文件夹路径
source_folder, parent_folder = opfiles.OpFiles.select_folder()
newfolder = source_folder.split("/")[-1] + "_切片理由2个或以上xdotx"
newfolder_path = os.path.normpath(os.path.join(parent_folder, newfolder))

try:
    os.makedirs(newfolder_path)
except FileExistsError:
    pass

articlecodes = []#"条文编号"
jsonnames = []#。json文件的文件名
row = 0
icount = 0
icount2 = 0
# 遍历源文件夹中的所有文件
bmod = True
for jsonname in os.listdir(source_folder):
    bmod = False
    #定义数组
    
    filenames = []#"文档名称"对应的值
    sliceuuids = []#"切片id"
    slicetexts = []#"切片不带格式"
    slicetext_format_list = []#"切片带格式"

    if jsonname.endswith('.json') or jsonname.endswith('.Json'):
        file_path = os.path.join(source_folder, jsonname)
        #jsonnames.append(jsonname)
        # 打开并读取JSON文件
        with open(file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
            icount2 = icount2 + 1
        try:
            for entry in data:  
                if "文档名称" in entry and "条文编号" in entry and "切片不带格式" in entry and "切片带格式" in entry:
                    filename = entry["文档名称"]
                    slicetext = entry["切片不带格式"]
                    slicetext_format = entry["切片带格式"]
                    articlecode = entry["条文编号"]

#条文说明JSON文件中 存在“切片不带格式”和“切片带格式”中除了行首的条文编号不同，其余部分的文字内容一样的这种情况的文件挑选出来
                    newsubstring = slicetext.lstrip()
                    
                    pattern = r'(\d+\.\d+(?:\.\d+)?)'
                    #r'[A-Za-z0-9]+\.[0-9]+\.[A-Za-z0-9]+'

                    # Extracting the matching parts
                    newsubstring = newsubstring[:12]
                    matches1 = '、'.join(re.findall(pattern, newsubstring))
                    matchlist = matches1.split('、')
                    if len(matchlist) > 1:
                        pass
                        jsonnames.append(jsonname)
                        articlecodes.append(articlecode)
                        break
                    slicetexts.append(newsubstring)

                else:
                    print(f'{jsonname} 字段不全缺；条文编号：{str(articlecode)}')#
            # 将修改后的数据写入新的json文件
            # if bmod:
            #     with open(file_path, 'w', encoding='utf-8') as file:
            #         json.dump(data, file, ensure_ascii=False , indent=4) 
            # else:
            #     pass              

        except json.JSONDecodeError:
            print(f'{jsonname} 该文件单独查是什么问题')

jsonnames

# 遍历源文件夹中的所有文件
# for filename in os.listdir(source_folder):
#     if filename.endswith('.json') or filename.endswith('.Json'):
#         file_path = os.path.join(source_folder, filename)
#         if filename in jsonnames:
#             pass
#             shutil.move(file_path,newfolder_path)


# 创建一个新的Excel工作簿和工作表
workbook = openpyxl.Workbook()
sheet = workbook.active


row = 0
for index, item in enumerate(jsonnames):

    row += 1
    sheet.cell(row=row, column=1, value=item)

    item2 = articlecodes[index]
    sheet.cell(row=row, column=2, value=item2)

excel_file_name = source_folder.split("/")[-1] + "_切片理由2个或以上xdotx.xlsx"
excelfolder_path = os.path.normpath(os.path.join(parent_folder, excel_file_name))
workbook.save(excelfolder_path)