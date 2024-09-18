#记录json里的文件名+改名

#改名
# "城市公共设施规划规范GB504422008",
# "电磁屏蔽室工程施工及质量验收规范GBT511032015",
# "沉井与气压沉箱施工规范"
# 成：
# 《城市公共设施规划规范》GB 50442-2008
# 《电磁屏蔽室工程施工及质量验收规范》GB T51103-2015
# 《沉井与气压沉箱施工规范》
import re
import os
import sys
from pathlib import Path
import json
import shutil
import openpyxl

# 获取当前文件的父目录
parent_dir = str(Path(__file__).resolve().parent.parent)
# 将父目录添加到sys.path
sys.path.append(parent_dir)

from filesfunction import opfiles

def format_standard_code_final(code):
    # 正则表达式匹配标准名称、标准代码和年份（允许标准代码中间有空格）
    pattern = r'^(.*?)([A-Z]+(?:\s*[A-Z]+)*)\s*(\d+)(?:(\d{4}))?$'
    match = re.search(pattern, code)

    if match:
        # 提取匹配的部分
        name, code_prefix, code_number, year = match.groups()
        # 去除标准代码中的空格
        code_prefix = code_prefix.replace(' ', '')
        # 处理标准代码和年份格式
        if year:
            formatted_code = f"《{name}》{code_prefix} {code_number}-{year}"
        else:
            formatted_code = f"《{name}》{code_prefix} {code_number}"
    else:
        # 如果没有匹配到标准代码和年份，则只添加书名号
        formatted_code = f"《{code}》"

    return formatted_code


def add_dash_to_last_four_digits(input_string):
    # 匹配字符串最后四个字符是数字的部分
    return re.sub(r'(\d{2})(\d{4})$', r'\1-\2', input_string)

# 定义文件夹路径
source_folder, parent_folder = opfiles.OpFiles.select_folder()

filenames = []
jsonnames = []
icount = 0
icount2 = 0
# 遍历源文件夹中的所有文件
for filename in os.listdir(source_folder):
    icount = icount + 1  
    if filename.endswith('.json') or filename.endswith('.Json'):
        file_path = os.path.join(source_folder, filename)
        jsonnames.append(filename)
        # 打开并读取JSON文件
        with open(file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
  
        try:
            for entry in data:
                jsonnames.append(1)
                icount2 = icount2 + 1    
                if "文档名称" in entry:
                    name = entry["文档名称"]
                    filenames.append(name)
                        # 移动文件到 "版本一" 文件夹
                        #shutil.move(file_path, target_folder)
                        #print(f'{filename} 已移动到 {target_folder}')
                        # if os.path.exists(file_path):
                        #     os.remove(file_path)
                        #     print(f"{file_path} 已成功删除")
                        # else:
                        #     print(f"文件 {file_path} 不存在")
                         
                    break
                else:
                    print("entry does not have a '版本' key")
            

        except json.JSONDecodeError:
            print(f'{filename} 不是有效的 JSON 文件')
    if not icount2 == icount:
        pass


# 创建一个新的Excel工作簿和工作表
workbook = openpyxl.Workbook()
sheet = workbook.active

row = 0
for item in filenames:
    row += 1
    sheet.cell(row=row, column=1, value=item)
    name_mod = format_standard_code_final(item)
    sheet.cell(row=row, column=3, value=name_mod)

    name_mod_year = add_dash_to_last_four_digits(name_mod)
    sheet.cell(row=row, column=5, value=name_mod_year)

excel_file_name = source_folder.split("/")[-1] + "_提取文件名.xlsx"
excelfolder_path = os.path.normpath(os.path.join(parent_folder, excel_file_name))
workbook.save(excelfolder_path)