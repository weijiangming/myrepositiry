#根据编号信息 判断库里是否已有此编号
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import os
import sys
from pathlib import Path
import re

# 获取当前文件的父目录
parent_dir = str(Path(__file__).resolve().parent.parent)
# 将父目录添加到sys.path
sys.path.append(parent_dir)

from filesfunction import opfiles

# 创建 Tkinter 主窗口（通常隐藏）
root = tk.Tk()
root.withdraw()

# 打开文件选择对话框
file_path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel文件', '*.xlsx')])

# 初始化一个空字典
yesorno_list = []
name_list = []
code_list = []

ck_index_list = []
ck_name_list = []
ck_code_list = []

def extract_pattern(text):
    # 正则表达式匹配规则：从字符串中提取符合“数字+"-"+"数字"”的模式，并要求字符串中的总长度不超过20个字符
    pattern = r'\b(\d+\.?\d*-\d{4})\b'
    match = re.search(pattern, text)
    if match:
        return match.group(1)
    return None

# 检查用户是否选择了文件
if file_path:
    # 加载选定的 Excel 文件
    workbook = load_workbook(filename=file_path)
    
    # 选择活动的工作表（假设是第一个表）
    sheet = workbook.active
    # 遍历每一行，将 A 列作为键，C 列作为值，假设第一行为表头，从第二行开始
    
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=5, max_col=6, values_only=False):
        cell = row[0]
        yesorno = cell.value
        name = row[1].value  # B列

        if cell.row == 8:
            pass

        yesorno_list.append(yesorno)
        name_list.append(name)

        if (name == "" or name is None) and yesorno == "有":
            sheet.cell(row=cell.row, column=7, value="检查")
        elif (name != "" and name != None) and yesorno == "无":
            sheet.cell(row=cell.row, column=7, value="检查")

        if name is None:
            pass

    #保存修改后的Excel文件
    workbook.save(file_path)
else:
    print("用户取消了选择文件")



