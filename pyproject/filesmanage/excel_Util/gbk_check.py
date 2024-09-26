#工标库检查（原测试用QA用例：简短专业词汇）
import re
import os
import sys
from pathlib import Path
import json
import openpyxl
import random

parent_dir = str(Path(__file__).resolve().parent.parent)# 获取当前文件的父目录
sys.path.append(parent_dir)# 将父目录添加到sys.path
from filesfunction import opfiles

#提取切片id、切片不带格式、切片带格式、条文编号行首 编号，判断是否有编号 又有漏对数量（切片id）。

#"文档id": "242de7ff-3c34-4dc2-9c26-be646b96993a",
# "切片id": "3f694971-b5fc-4019-a57f-7ac47b127625",
# "文档名称": "《现浇改性石膏墙体应用技术规程》T/CECS 971—2021；T/CREA 008—2021 ",
# "所属章节标题": "1 总 则",
# "切片不带格式": "1.0.2 本规程适用于抗震设防烈度8度及以下地区，一般工业与民用建筑中采用现浇改性石膏墙体作为内隔墙的设计、施工及验收。",
# "切片带格式": "1.0.2 本规程适用于抗震设防烈度8度及以下地区，一般工业与民用建筑中采用现浇改性石膏墙体作为内隔墙的设计、施工及验收。",
# "条文编号": "1.0.2",
# "版本": "1",

# 定义文件夹路径
source_folder, parent_folder = opfiles.OpFiles.select_folder()

# 创建一个新的Excel工作簿和工作表
workbook = openpyxl.Workbook()
sheet = workbook.active

row = 0
icount = 0
icount2 = 0
# 遍历源文件夹中的所有文件
for jsonname in os.listdir(source_folder):
    icount = icount + 1  

    #定义数组
    jsonnames = []#。json文件的文件名
    filenames = []#"文档名称"对应的值
    sliceuuids = []#"切片id"
    slicetexts = []#"切片不带格式"
    slicetext_format_list = []#"切片带格式"
    articlecodes = []#"条文编号"
    if jsonname.endswith('.json') or jsonname.endswith('.Json'):
        file_path = os.path.join(source_folder, jsonname)
        #jsonnames.append(jsonname)
        # 打开并读取JSON文件
        with open(file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
            icount2 = icount2 + 1
        try:
            for entry in data:  
                if "文档名称" in entry and "切片id" in entry and "条文编号" in entry and "切片不带格式" in entry and "切片带格式" in entry:
                    filename = entry["文档名称"]
                    sliceuuid = entry["切片id"]
                    slicetext = entry["切片不带格式"]
                    slicetext_format = entry["切片带格式"]
                    articlecode = entry["条文编号"]

                    if 0 == len(filename) or 0 == len(sliceuuid) or 0 == len(slicetext) or 0 == len(slicetext_format) or 0 == len(str(articlecode)) : 
                        print(f'{jsonname} 五字段中有值为空')
                        break
                    else:
                        jsonnames.append(jsonname)
                        filenames.append(filename)
                        sliceuuids.append(sliceuuid)
                        slicetexts.append(slicetext)
                        slicetext_format_list.append(slicetext_format)
                        articlecodes.append(articlecode)
                        
                else:
                    print(f'{jsonname} 五字段不全缺')#
                    break

        except json.JSONDecodeError:
            print(f'{jsonname} 该文件单独查是什么问题')
    if not icount == icount2:
        print(f'{jsonname} 该文件单独查是什么问题')

    #分析5个字段有什么问题

    for index, item in enumerate(sliceuuids):
        jsonname = jsonnames[index]
        filename = filenames[index]
        sliceuuid = sliceuuids[index]
        slicetext = slicetexts[index]
        slicetext_format = slicetext_format_list[index]
        articlecode = articlecodes[index]

        pattern2 = r'[A-Za-z0-9]\.\d+\.\d+'
        pattern = r'[A-Za-z0-9]+\.[0-9]+\.[A-Za-z0-9]+'
        matches1 = re.findall(pattern, slicetext)
        slicetextres = ""
        slicetext_formatres = ""
        if matches1:
            slicetextres = matches1[0]

        matches1 = re.findall(pattern, slicetext_format)
        if matches1:
            slicetext_formatres = matches1[0]
        
        if articlecode == slicetextres and articlecode == slicetext_formatres:#用于发现问题
        #if articlecode in slicetextres and articlecode in slicetext_formatres:
            pass
        else:
            print(f'{jsonname} 条文编号{str(articlecode)}:切片不带格式、切片带格式、条文编号行首的标号不一致')#
            jsonname
            filename
            articlecode
            slicetext
            slicetextres
            slicetext_format
            slicetext_formatres
            row += 1
            sheet.cell(row=row, column=1, value=jsonname)
            sheet.cell(row=row, column=2, value=filename)
            sheet.cell(row=row, column=3, value=articlecode)
            sheet.cell(row=row, column=4, value=slicetext)
            sheet.cell(row=row, column=5, value=slicetextres)
            sheet.cell(row=row, column=6, value=slicetext_format)
            sheet.cell(row=row, column=7, value=slicetext_formatres)
            #continue
            break
pass

excel_file_name = source_folder.split("/")[-1] + "_提取文件名.xlsx"
excelfolder_path = os.path.normpath(os.path.join(parent_folder, excel_file_name))
workbook.save(excelfolder_path)




# if len(onearticles) == len(articlecodes) and len(onearticles) == len(filenames):
#     #规范名称 1 如工：
#     row = 0
#     for index, item in enumerate(onearticles):
#         row += 1
#         pattern = r"[\u4e00-\u9fff]+"
#         match2 = re.findall(pattern, item)
#         result2 = ""
#         if match2:
#             result2 = match2[0]

#         sheet.cell(row=row, column=1, value=result2)
        
#     #答案 2  如：1.01
#     row = 0

#     for item in articlecodes:
#         row += 1
#         sheet.cell(row=row, column=2, value=item)

#     #截图或内容分片 3
#     row = 0
#     for item in articles:
#         row += 1
#         sheet.cell(row=row, column=3, value=item)

#     #规范名称及编号 4 如：《工程结构通用规范》GB55001-2021
#     row = 0
#     # for item in filenames:
#     #     row += 1
#     #     sheet.cell(row=row, column=4, value=item)
#     for item in jsonnames:
#         row += 1
#         sheet.cell(row=row, column=4, value=item)

#     excel_file_name = source_folder.split("/")[-1] + "_提取文件名.xlsx"
#     excelfolder_path = os.path.normpath(os.path.join(parent_folder, excel_file_name))
#     workbook.save(excelfolder_path)