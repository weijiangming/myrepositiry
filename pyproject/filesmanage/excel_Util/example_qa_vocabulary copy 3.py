##测试用QA用例：简短专业词汇
import re
import os
import sys
from pathlib import Path
import json
import openpyxl
import random


# "所属章节标题": "2  术语和符号\n2.1  术    语",
# "切片不带格式": "2.1.6 粮食密度 mass density of grain\n单位体积的粮食质量。\n2.1.7 粮食重度 gravity density of grain\n单位体积粮食所受的重力。\n2.1.8 粮食压力 pressure of grain\n粮食作用在接触物体表面上的力。\n2.1.9 原粮 raw grain\n未经加工的谷物、豆类及薯类的总称。\n2.1.10 成品粮 product grain\n原粮经过机械等方式加工后形成的产品。",
# "条文编号": "2.1.6、2.1.7、2.1.8、2.1.9、2.1.10",



# 获取当前文件的父目录
parent_dir = str(Path(__file__).resolve().parent.parent)
# 将父目录添加到sys.path
sys.path.append(parent_dir)

from filesfunction import opfiles

def extract_number(text):
    """
    提取文本中的数字编号和术语对应关系，并返回一个字典。

    Args:
      text: 包含数字编号和术语的文本

    Returns:
      包含数字编号和术语的字典。如果没有找到则返回空字典。
    """

    # 将文本按照行分割
    lines = text.split('\n')

    # 创建一个字典，用于存储数字编号和术语的对应关系
    term_dict = {}
    for line in lines:
        parts = line.split(' ', maxsplit=1)  # 只分割一次，避免多余的空格
        if len(parts) == 2:
            number, term = parts
            term_dict[term.strip()] = number.strip()

    return term_dict

# 定义文件夹路径
source_folder, parent_folder = opfiles.OpFiles.select_folder()

filenames = []
jsonnames = []
articlecodes = []
articles = []
onearticles = []
icount = 0
icount2 = 0
# 遍历源文件夹中的所有文件
for filename in os.listdir(source_folder):
    icount = icount + 1  
    if filename.endswith('.json') or filename.endswith('.Json'):
        file_path = os.path.join(source_folder, filename)
        #jsonnames.append(filename)
        # 打开并读取JSON文件
       
        with open(file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
            try:
                filenames_two = []
                jsonnames_two = []
                articlecodes_two = []
                articles_two = []
                onearticles_two = []

                iCountT = 0
                for entry in data:  
                    iCountT = iCountT + 1
                    if "文档名称" in entry and "条文编号" in entry and "切片不带格式" in entry and "所属章节标题" in entry:
                        name = entry["文档名称"]
                        article = entry["切片不带格式"]
                        articlecode = entry["条文编号"]
                        titil = entry["所属章节标题"]
                        #if "1.0.1" in str(articlecode):
                        #if len(str(articlecode)) < 9 and len(str(articlecode)) > 4 and "术" in titil and "语" in titil and "2.1." in articlecode:
                        if len(str(articlecode)) > 4 and "术" in titil and "语" in titil:
                            term_dict = {}
                            term_dict = extract_number(article)
                            for term, number in term_dict.items():
                                if not article in articles_two:
                                    articlecodes_two.append(number)
                                    onearticles_two.append(term)
                                    articles_two.append(article)
                                    filenames_two.append(name)
                                    jsonnames_two.append(filename)
                                #break
                            #break
                    else:
                        print("entry does not have a '版本' key")
                if len(articlecodes_two) > 0:
                    element_count = len(articlecodes_two)
                    random_number = random.randint(0, element_count-1)
                    if random_number >= 0 and random_number < len(articlecodes_two):
                        articlecodes.append(articlecodes_two[random_number])
                        onearticles.append(onearticles_two[random_number])
                        articles.append(articles_two[random_number])
                        filenames.append(filenames_two[random_number])
                        jsonnames.append(jsonnames_two[random_number])
    
            except json.JSONDecodeError:
                print(f'{filename} 不是有效的 JSON 文件')
    if not icount2 == icount:
        pass

# 创建一个新的Excel工作簿和工作表
workbook = openpyxl.Workbook()
sheet = workbook.active

if len(onearticles) == len(articlecodes) and len(onearticles) == len(filenames):
    #规范名称 1 如工：
    row = 0
    for index, item in enumerate(onearticles):
        row += 1
        pattern = r"[\u4e00-\u9fff]+"
        match2 = re.findall(pattern, item)
        result2 = ""
        if match2:
            result2 = match2[0]

        sheet.cell(row=row, column=1, value=result2)
        
    #答案 2  如：1.01
    row = 0

    for item in articlecodes:
        row += 1
        sheet.cell(row=row, column=2, value=item)

    #截图或内容分片 3
    row = 0
    for item in articles:
        row += 1
        sheet.cell(row=row, column=3, value=item)

    #规范名称及编号 4 如：《工程结构通用规范》GB55001-2021
    row = 0
    for item in filenames:
        row += 1
        sheet.cell(row=row, column=4, value=item)
    
    row = 0
    for item in jsonnames:
        row += 1
        sheet.cell(row=row, column=5, value=item)

    excel_file_name = source_folder.split("/")[-1] + "_提取文件名.xlsx"
    excelfolder_path = os.path.normpath(os.path.join(parent_folder, excel_file_name))
    workbook.save(excelfolder_path)