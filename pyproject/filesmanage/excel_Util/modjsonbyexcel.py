#根据文件名对应的标准格式规范名修改json的”"文档名称"的值
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import os
import sys
from pathlib import Path
import json

# 获取当前文件的父目录
parent_dir = str(Path(__file__).resolve().parent.parent)
# 将父目录添加到sys.path
sys.path.append(parent_dir)

from filesfunction import opfiles

# 创建 Tkinter 主窗口（通常隐藏）
root = tk.Tk()
root.withdraw()

# 打开文件选择对话框
file_path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel文件', '*.xlsx')])

source_folder, parent_folder = opfiles.OpFiles.select_folder()

newfolder = source_folder.split("/")[-1] + "_newname"
newfolder_path = os.path.normpath(os.path.join(parent_folder, newfolder))

# try:
#     os.makedirs(newfolder_path)
# except FileExistsError:
#     pass

# 初始化一个空字典
data_dict = {}

# 检查用户是否选择了文件
if file_path:
    # 加载选定的 Excel 文件
    workbook = load_workbook(filename=file_path)
    
    # 选择活动的工作表（假设是第一个表）
    sheet = workbook.active
    # 遍历每一行，将 A 列作为键，C 列作为值，假设第一行为表头，从第二行开始
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3, values_only=True):
        key = row[0]  # A列
        value = row[1]  # B列
        data_dict[key] = value

else:
    print("用户取消了选择文件")

# 遍历源文件夹中的所有文件
for filename in os.listdir(source_folder):
    if filename.endswith('.json') or filename.endswith('.Json'):
        file_path = os.path.join(source_folder, filename)
  
        if filename in data_dict:
            newname = data_dict[filename]

            with open(file_path, 'r', encoding='utf-8') as json_file:
                data = json.load(json_file)
    
            try:
                #修改 "文档名称"
                for item in data:
                    if "文档名称" in item:
                        item['文档名称'] = newname

                # 将修改后的数据写入新的json文件
                with open(file_path, 'w', encoding='utf-8') as file:
                    json.dump(data, file, ensure_ascii=False , indent=4)        
                            
            except json.JSONDecodeError:
                print(f'{filename} 不是有效的 JSON 文件')


