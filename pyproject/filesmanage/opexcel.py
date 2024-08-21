import os
import openpyxl
import time
import tkinter as tk
from tkinter import filedialog

start_time = time.time()

#定义常数
basepath = "F:\资料分类"
sheet_name='分类文件夹名'
exl_name="资料分类.xlsx"

#该文档的功能是分类文件夹名记录在F:\资料分类\资料分类.xlsx里的工作表'分类文件夹名' 和建立“资料分类”文件夹以及子文件夹
def is_merged_cell(cell, merged_ranges):
    for merged_cell_range in merged_ranges:
        if cell.coordinate in merged_cell_range:
            return True
    return False

def write_dict_to_excel(data_dict, excel_file, sheet_name):
    """
    将字典数据写入 Excel

    Args:
        data_dict: 字典数据
        excel_file: Excel 文件名
        sheet_name: 工作表名称
    """
    try:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook[sheet_name]
        worksheet.delete_rows(1, worksheet.max_row)  # 清空工作表内容
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

    for fdrname, fdrnamefullpath in data_dict.items():
        worksheet.append([fdrname, fdrnamefullpath])

    workbook.save(excel_file)

def select_folder():
    """
    选择文件夹并返回其路径和上级目录路径。

    Returns:
        tuple: (selected_folder, parent_folder)
    """
    root = tk.Tk()
    root.withdraw()
    selected_folder = filedialog.askdirectory()

    if selected_folder:
        # 获取上级目录
        parent_folder = os.path.dirname(selected_folder)
        return selected_folder, parent_folder
    else:
        return None, None


print('当前工作目录 :')
print(os.getcwd(), '\n')

wb = openpyxl.load_workbook('C:\\Users\\jimee\\Desktop\\房建类施工方案清单.xlsx')
print('wb 类型 :')
print(type(wb), '\n')

# 选择工作表
sheet = wb['房建类方案标签整理']  # 根据需要替换为实际的工作表名称
#sheet = wb.active
print('表名 - ' + sheet.title, '\n')
rowsheet = sheet.max_row

selected_folder, parent_folder = select_folder()

if selected_folder:
    print(f"您选择的文件夹是：{selected_folder}")
    print(f"其上级目录是：{parent_folder}")
else:
    print("您取消了选择。")


# 获取工作表中所有合并单元格的范围
merged_ranges = sheet.merged_cells.ranges
merged_ranges_count = len(merged_ranges)

print(f"合并单元格总个数有: {merged_ranges_count}个。")

fullpath2foldername = {}


icount = 0
# 获取合并单元格的范围
for merged_cell_range in sheet.merged_cells.ranges:
    # 获取合并单元格的起始和结束单元格
    min_col, min_row, max_col, max_row = merged_cell_range.bounds
    #第2列的合并单元格
    if min_col == 2:
      # 获取合并单元格的左上角单元格
        start_cell = sheet.cell(row=min_row, column=min_col)
        folder_name = start_cell.value
        #print(f"Start cell value: {start_cell.value}")
        #print(f"Merged cell range: {merged_cell_range}")
        icount += 1

        #建第二列的文件夹
        new_path = os.path.join(basepath, folder_name)
        if not os.path.exists(new_path):
            os.makedirs(new_path)
            #print(f"一级目录{new_path} 创建成功")
            #print(new_path)
        #else:
            #print(f"一级目录{new_path}已存在，无需创建")
            
            
        #建第三列的文件夹
        #开始处理第三列的内容，判断是第二列子项
        #1.合并单元格
        for merged_cell_range in sheet.merged_cells.ranges:
            min_colT, min_rowT, max_colT, max_rowT = merged_cell_range.bounds
            #选取第3列的合并单元格
            if min_colT == 3:
                start_cellT = sheet.cell(row=min_rowT, column=min_colT)
                #选取被第二列包含的
                if (min_rowT >= min_row) and (min_rowT <= max_row):
                    folder_nameT = start_cellT.value
                    new_pathT = os.path.join(new_path, folder_nameT)
                    fullpath2foldername[folder_nameT] = new_pathT
                    if not os.path.exists(new_pathT):
                        os.makedirs(new_pathT)
                        #print(f"子目录{new_pathT} 创建成功")
                        print(new_pathT)
                    else:
                        #print(f"子目录{new_pathT}已存在，无需创建")
                        print(new_pathT)


        #2.第三列的单个单元格
        col = 3
        for row in range(1, rowsheet + 1):
            cell = sheet.cell(row=row, column=col)
            #2.1排除在合并单元格里的
            if (row >= min_row) and (row <= max_row):
            #if not cell.is_merged:
                 if not is_merged_cell(cell, merged_ranges):
                    folder_nameT = cell.value
                    new_pathT = os.path.join(new_path, folder_nameT)
                    fullpath2foldername[folder_nameT] = new_pathT
                    if not os.path.exists(new_pathT):
                        os.makedirs(new_pathT)
                        #print(f"子目录单单元格{new_pathT}创建成功")
                        print(new_pathT)
                    else:
                        #print(f"子目录单单元格{new_pathT}已存在，无需创建")
                        print(new_pathT)


    #break
    """折叠
    print(f"Start cell: ({min_row}, {min_col})")
    print(f"End cell: ({max_row}, {max_col})")

   

    # 获取合并单元格的行和列
    cell_row = start_cell.row
    cell_column = start_cell.column

    # 获取合并单元格的字体信息
    cell_font = start_cell.font

    # 获取合并单元格的填充信息
    cell_fill = start_cell.fill

    # 打印单元格信息
    print(f"Cell value: {start_cell.value}")
    #print(f"Cell row: {cell_row}")
    #print(f"Cell column: {cell_column}")
    #print(f"Cell font: {cell_font}")
    #print(f"Cell fill: {cell_fill}")
    """
print(f"第2列的合并单元格个数： {icount}")

#for fdrname, fdrnamefullpath in fullpath2foldername.items():

#
excel_file = os.path.join(basepath, exl_name)
write_dict_to_excel(fullpath2foldername, excel_file, sheet_name)

end_time = time.time()
print("程序运行时间：%s 秒" % (end_time - start_time))