import os
import openpyxl
import shutil
import win32com.client
from filesfunction import opfiles

def is_merged_cell(cell, merged_ranges):
    for merged_cell_range in merged_ranges:
        if cell.coordinate in merged_cell_range:
            return True
    return False



wb = openpyxl.load_workbook('C:\\Users\\admin\\Desktop\\构力JSON_正文_v2.2.xlsx')
print('wb 类型 :')
print(type(wb), '\n')

# 选择工作表
sheet = wb['Sheet']
maxrow_sheetkey = sheet.max_row
print('表名 - ' + sheet.title, '\n')

selected_folder, parent_folder = opfiles.OpFiles.select_folder()

des_path = os.path.normpath(os.path.join(parent_folder, "缺少的"))
try:
    os.makedirs(des_path)
except FileExistsError:
    pass


filenames = []
colB = 2
colA = 1
for row in range(1, maxrow_sheetkey + 1):
    cell = sheet.cell(row=row, column=colB) 
    search_term = cell.value
    #file_Nosuf = opfiles.OpFiles.remove_suffix(search_term, 5)
    if search_term == "M":
        cellA = sheet.cell(row=row, column=colA)
        valueA = cellA.value
        filenames.append(valueA)


for root, dirs, files in os.walk(selected_folder):
    #文件
    if root == selected_folder:
        for file in files:
            if file.endswith('.json'):
                if file in filenames:
                    file
                    file_path = os.path.normpath(os.path.join(selected_folder, file))
                    try:
                        shutil.move(file_path,des_path)
                    except shutil.Error as e:
                        print(file_path+"未成功移动!")  
                    
