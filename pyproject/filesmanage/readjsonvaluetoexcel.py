import os
import json
import openpyxl
from filesfunction import opfiles

# 创建一个新的Excel工作簿和工作表
workbook = openpyxl.Workbook()
sheet = workbook.active


selected_folder, parent_folder = opfiles.OpFiles.select_folder()
for root, dirs, files in os.walk(selected_folder):
    if root == selected_folder:
        for dir in dirs:
            pass
            file_path = os.path.normpath(os.path.join(selected_folder, dir))
            for root2, dirs2, files2 in os.walk(file_path):
                if root2 == file_path:
                    for dir2 in dirs2:
                        if dir2== "条文说明":
                            continue


                        row = 1
                        file_path2 = os.path.normpath(os.path.join(file_path, dir2))
                        for root3, dirs3, files3 in os.walk(file_path2):

                            #读取JSON文件
                            labelValues = []
                            for file in files3:
                                if not file.endswith('.json'):
                                    continue
                                filepath = os.path.normpath(os.path.join(root3, file))
                                with open(filepath, 'r', encoding='utf-8') as f:
                                    data_dict = json.load(f)

                                try:
                                        # 提取所有 "label" 值
                                    # 判断是否存在 data 和 entryContents
                                    if 'data' in data_dict:
                                        if 'entryContents' in data_dict['data']:


                                            if 'data' in data_dict and data_dict['data'] is not None and 'entryContents' in data_dict['data'] and data_dict['data']['entryContents'] is not None:

                                                for entry in data_dict['data']['entryContents']:
                                                    if 'label' in entry:
                                                        labelValues.append(entry['label'])
                                                    else:
                                                        print("entry does not have a 'label' key")
                                            else:
                                                print("data or entryContents is missing")

                                except json.JSONDecodeError as e:
                                    print(f"JSON解析错误：{e}, 文件：{filepath}")
                                    continue
                                except KeyError as e:
                                    print(f"KeyError: {e}, 文件：{filepath}")
                                    continue
                                pass
                            row += 1
                        sheet.cell(row=row, column=2, value=dir2)
                        row += 1
                        # 将字符写入Excel的C列
                        for lablevalue in labelValues:
                            sheet.cell(row=row, column=2, value=lablevalue)
                            row += 1
                        # 保存Excel文件
                        excelfolder_path = os.path.normpath(os.path.join(file_path, dir2 + ".xlsx"))
                        workbook.save(excelfolder_path)
                        break
                        