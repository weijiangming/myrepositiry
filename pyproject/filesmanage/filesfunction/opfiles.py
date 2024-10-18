import os
import tkinter as tk
from tkinter import filedialog
import openpyxl


class OpFiles:

    @staticmethod
    def select_folder():
        """
        选择文件夹并返回其路径和上级目录路径。

        Returns:
            tuple: (selected_folder, parent_folder)
        """
        root = tk.Tk()
        root.withdraw()
        selected_folder = filedialog.askdirectory()

        if selected_folder:
            # 获取上级目录
            parent_folder = os.path.dirname(selected_folder)
            return selected_folder, parent_folder
        else:
            return None, None
        
    @staticmethod
    def select_excel_file():
        """
        选择Excel文件并返回其路径和上级目录路径。

        Returns:
        tuple: (selected_file, parent_folder)
        """

        root = tk.Tk()
        root.withdraw()

        # 设置文件类型过滤器，只允许选择Excel文件
        filetypes = [('Excel文件', '*.xlsx'), ('Excel文件', '*.xls')]
        selected_file = filedialog.askopenfilename(filetypes=filetypes)

        if selected_file:
            return selected_file
        else:
            return None
        

    @staticmethod
    def remove_suffix(filename, num):
        """
        从文件名中去除  后缀

        Args:
            filename: 包含后缀的文件名
            num:去除文件名后面位数

        Returns:
            去除 后缀后的文件名
        """

        # 使用字符串切片，直接截取到 .rar 前面的部分
        return filename[:-num]

    @staticmethod
    def write_dict_to_excel(data_dict, excel_file, sheet_name):
        """
        将字典数据写入 Excel

        Args:
            data_dict: 字典数据
            excel_file: Excel 文件名
            sheet_name: 工作表名称
        """
        try:
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook[sheet_name]
            worksheet.delete_rows(1, worksheet.max_row)  # 清空工作表内容
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name

        for fdrname, fdrnamefullpath in data_dict.items():
            worksheet.append([fdrname, fdrnamefullpath])

        workbook.save(excel_file)

    @staticmethod
    def write_1d_list_to_excel(data_list, excel_file, sheet_name, start_row=1, start_column=1):
        """
        将一维列表数据写入 Excel 的一行中

        Args:
            data_list: 一维列表数据
            excel_file: Excel 文件名
            sheet_name: 工作表名称
            start_row: 开始写入的行号
            start_column: 开始写入的列号
        """

        if not isinstance(data_list, list):
            raise TypeError("data_list must be a list")

        try:
            workbook = openpyxl.load_workbook(excel_file)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        try:
            worksheet = workbook[sheet_name]
        except KeyError:
            worksheet = workbook.create_sheet(sheet_name)

        # 从指定的起始列开始写入数据，仅写入一行
        for col_index, cell_value in enumerate(data_list, start=start_column):
            worksheet.cell(row=start_row, column=col_index, value=cell_value)

        try:
            workbook.save(excel_file)
        except PermissionError:
            print(f"无法保存文件 {excel_file}: 请检查文件是否已打开或权限是否足够")



