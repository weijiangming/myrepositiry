import os
import openpyxl
import shutil
import win32com.client

#注意检查61~72行的常量是否有变动
#该代码功能是按分类将文件移动到对应的文件夹里
#文件分类存放到文件夹
#文件没分类记录到未分类.txt
#1.目录列表
#2.获取搜索词
""""3.目录列表关联搜索词 用三个字典搞定 key2secondnameDic (单搜索词：二级名)；
    2) secondname2firstlevelDic (二级名：一级目录)；3) secondname2secondlevel (二级名：二级目录)。"""
#4.遍历basepath文件夹

def is_merged_cell(cell, merged_ranges):
    for merged_cell_range in merged_ranges:
        if cell.coordinate in merged_cell_range:
            return True
    return False

def get_folder_names(root_dir):
    """
    Args:
        root_dir:指定根目录路径
        Returns：指定目录下所有文件夹名称的列表
    """
    folder_names= []
    for root, dirs, files in os.walk(root_dir):
        for dir in dirs:
            folder_names.append(dir)
    return folder_names

def get_full_folder_paths(root_dir):
    """
    Args:
        root_dir: 指定根目录路径
    Returns:
        指定目录下所有文件夹的全路径名的列表
    """
    folder_paths = []
    for root, dirs, _ in os.walk(root_dir):
        for dir in dirs:
            full_path = os.path.join(root, dir)
            folder_paths.append(full_path)
    return folder_paths

def get_folder2fullpaths(root_dir):
    """
    Args:
        root_dir: 指定根目录路径
    Returns:
        指定目录下所有文件夹的全路径名的列表
    """
    folder2fullpaths = {}
    for root, dirs, _ in os.walk(root_dir):
        for dir in dirs:
            full_path = os.path.join(root, dir)
            folder2fullpaths[dir] = full_path
    return folder2fullpaths

basepath = "F:\文档\电器"

destpath = "F:\资料分类"
exl_name="资料分类.xlsx"
sheet_name='分类文件夹名'

print('当前工作目录 :')
print(os.getcwd(), '\n')

wb = openpyxl.load_workbook('C:\\Users\\jimee\\Desktop\\房建类施工方案清单.xlsx')
print('wb 类型 :')
print(type(wb), '\n')

# 选择工作表
sheet = wb['房建类方案标签整理']
maxrow_sheetkey = sheet.max_row
print('表名 - ' + sheet.title, '\n')

# 获取工作表中所有合并单元格的范围
merged_ranges = sheet.merged_cells.ranges

#1.目录列表
dataClassifyxlsx = os.path.join(destpath, exl_name)
wb_classify = openpyxl.load_workbook(dataClassifyxlsx)
sheet2 = wb_classify[sheet_name]
maxrow_sheetfolder = sheet2.max_row
col = 1
col2 = 2
secondname2secondlevel = {}
for row in range(1, maxrow_sheetfolder + 1):
    cell = sheet2.cell(row=row, column=col)
    cell2 = sheet2.cell(row=row, column=col2)
    secondname2secondlevel[cell.value] = cell2.value

#2.获取搜索词 每一个单以搜索词对应一个二级目录名存在key2secondnameDic里
data = []
key2secondnameDic = {} #(单搜索词：二级名)；
#2.1合并单元格
for merged_cell_range in sheet.merged_cells.ranges:
    min_col, min_row, max_col, max_row = merged_cell_range.bounds
    #选取第4列的合并单元格
    search_termlist = []
    if min_col == 4:
        start_cell = sheet.cell(row=min_row, column=min_col)
        search_term = start_cell.value
        #data.append(search_term)
        search_termlist = [item.strip() for item in search_term.split('、')]
        #print(search_termlist)
        data.append(search_termlist)

        # 二级文件夹分合并单元格和单元格两种情况处理
        #合并单元格
        foldernameList = []
        for merged_cell_range in sheet.merged_cells.ranges:
            min_colT, min_rowT, max_colT, max_rowT = merged_cell_range.bounds
            if min_colT == 3:
                if (min_rowT >= min_row) and (min_rowT <= max_row):
                    start_cellT = sheet.cell(row=min_rowT, column=min_colT)
                    folder_nameT = start_cellT.value
                    foldernameList.append(folder_nameT)
        
        #单个单元格
        colT = 3
        for row in range(1, maxrow_sheetkey + 1):
            cell = sheet.cell(row=row, column=colT)
            if not is_merged_cell(cell, merged_ranges):#不在合并单元格里边
                if(row >= min_row) and (row <= max_row):
                    folder_nameT = cell.value
                    foldernameList.append(folder_nameT)

        #单个搜索词匹配一个二级目录名
        for term in search_termlist:
            bfindfolder = False
            for foldername in foldernameList:
                if term in foldername:
                    bfindfolder = True
                    key2secondnameDic[term] = foldername
                    #if(len(foldernameList) != 1):
                    break
            if not bfindfolder:
                if len(foldernameList) > 0:
                    key2secondnameDic[term] = foldernameList[0]

#2.2第4列的单个单元格
col = 4
for row in range(2, maxrow_sheetkey + 1):
    cell = sheet.cell(row=row, column=col) 
    search_termlist = []
    #2.1排除在合并单元格里的
    if not is_merged_cell(cell, merged_ranges):
        search_term = cell.value
        #data.append(search_term)
        search_termlist = [item.strip() for item in search_term.split('、')]
        data.append(search_termlist)

        # 二级文件夹分合并单元格和单元格两种情况处理
        foldernameList = []
        """以下情况不存在
        #合并单元格
       
        for merged_cell_range in sheet.merged_cells.ranges:
            min_colT, min_rowT, max_colT, max_rowT = merged_cell_range.bounds
            if min_colT == 3:
                if (min_rowT == row):
                    start_cellT = sheet.cell(row=min_rowT, column=min_colT)
                    folder_nameT = start_cellT.value
                    foldernameList.append(folder_nameT)
        """
        #单个单元格
        colT = 3
        for rowT in range(1, maxrow_sheetkey + 1):
            cell = sheet.cell(row=rowT, column=colT)
            if not is_merged_cell(cell, merged_ranges):#不在合并单元格里边
                if(rowT == row):
                    folder_nameT = cell.value
                    foldernameList.append(folder_nameT)
                    
        #单个搜索词匹配一个二级目录名
        for term in search_termlist:
            bfindfolder = False
            for foldername in foldernameList:
                if term in foldername:
                    bfindfolder = True
                    key2secondnameDic[term] = foldername
                    #if(len(foldernameList) != 1):
                    break
            if not bfindfolder:
                if len(foldernameList) > 0:
                    key2secondnameDic[term] = foldernameList[0]


#3.目录列表关联搜索词 用两个字典搞定
#搜索词对应文件夹的字典
#print(key2secondnameDic)

#4.所有文件夹名
folder2fullpaths = get_folder2fullpaths(basepath)
#print(all_folders)

 # 启动 Word 应用程序
word = win32com.client.Dispatch("Word.Application")

# try:
    #foldernameT文件夹名；folderfullnameT文件夹名带路径；phrase搜索词；plan二级名;pathfull目的地文件夹
for foldernameT, folderfullnameT in folder2fullpaths.items():
    for phrase, plan in key2secondnameDic.items():
        if phrase in foldernameT:
            if plan in secondname2secondlevel:
                pathfull = secondname2secondlevel[plan]
                if os.path.exists(pathfull) and os.path.exists(folderfullnameT):#目录存在
                    if os.path.isdir(pathfull) and os.path.isdir(folderfullnameT):#是否是文件夹
                        try:
                            shutil.move(folderfullnameT,pathfull)
                        except shutil.Error as e:
                            print({folderfullnameT}+"未成功移动!") 

                            # # 遍历folderfullnameT文件夹下的文件
                            # for root, dirs, files in os.walk(folderfullnameT):
                            #     if root == folderfullnameT:  # 只处理指定文件夹下的这一层
                            #         isValid_docordocx = False
                            #         is_bpdf = False
                            #         isValid_size = False
                            #         has_ocordocx = False
                            #         page_count = 0
                            #         file_size = 0
                            #         for file in files:
                            #             file_full_path = os.path.normpath(os.path.join(root, file))
                            #             file_size = os.path.getsize(file_full_path) / 1024 /1024
                            #             if not isValid_size:
                            #                 if file_size < 1.5:
                            #                     continue
                            #                 else:
                            #                     isValid_size = True
                            #             if file.endswith('.docx') or file.endswith('.doc'):
                            #                 has_ocordocx = True
                            #                 try:
                            #                     # 打开指定的文档
                            #                     doc = word.Documents.Open(file_full_path)
                            #                     # 获取文档的页数
                            #                     page_count = doc.ComputeStatistics(2)  # 2代表页数
                            #                     # 关闭文档
                            #                     doc.Close()
                            #                     if page_count >= 50:
                            #                         isValid_docordocx = True
                            #                         break

                            #                     # print(f"文件: {file_full_path}")
                            #                     # print(f"大小: {file_size:.2f} MB")
                            #                     # print(f"页数: {page_count} 页")
                            #                     # print("=" * 25)
                            #                 except Exception as e:
                            #                     print(f"处理文件 {file_full_path} 时出错: {e}")
                            #                     if 'doc' in locals():
                            #                         doc.Close()

                            #             if file.endswith('.pdf'):
                            #                 isValid_pdf = True
                            #                 break

                            #         if isValid_docordocx:
                            #             try:
                            #                 shutil.move(folderfullnameT,pathfull)
                            #             except shutil.Error as e:
                            #                 print({folderfullnameT}+"未成功移动!") 
                            #         elif isValid_size:
                            #             try:
                            #                 shutil.move(folderfullnameT,pathfull)
                            #             except shutil.Error as e:
                            #                 print({folderfullnameT}+"未成功移动!") 
                            #         else:
                            #             print(f"文件过小：页数：{int(page_count)}页，大小：{file_size:.2f}MB，路径：{folderfullnameT}")                    
# finally:
#     # 退出 Word 应用程序
#     word.Quit()        

#secondname2secondlevel
""""key2secondnameDic (单搜索词：二级名)；
    secondname2firstlevelDic (二级名：一级目录)；
    secondname2secondlevel (二级名：二级目录)。"""
