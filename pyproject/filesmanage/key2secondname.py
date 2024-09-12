#关键词和文件分类对应
import os
import openpyxl
import shutil

#文件没分类记录到未分类.txt
#1.目录列表
#2.获取搜索词
""""3.目录列表关联搜索词 用三个字典搞定 key2secondnameDic (单搜索词：二级名)；
    2) secondname2firstlevelDic (二级名：一级目录)；3) secondname2secondlevel (二级名：二级目录)。"""
#4.遍历basepath文件夹

def is_merged_cell(cell, merged_ranges):
    for merged_cell_range in merged_ranges:
        if cell.coordinate in merged_cell_range:
            return True
    return False

def get_folder_names(root_dir):
    """
    Args:
        root_dir:指定根目录路径
        Returns：指定目录下所有文件夹名称的列表
    """
    folder_names= []
    for root, dirs, files in os.walk(root_dir):
        for dir in dirs:
            folder_names.append(dir)
    return folder_names

def get_full_folder_paths(root_dir):
    """
    Args:
        root_dir: 指定根目录路径
    Returns:
        指定目录下所有文件夹的全路径名的列表
    """
    folder_paths = []
    for root, dirs, _ in os.walk(root_dir):
        for dir in dirs:
            full_path = os.path.join(root, dir)
            folder_paths.append(full_path)
    return folder_paths

def get_folder2fullpaths(root_dir):
    """
    Args:
        root_dir: 指定根目录路径
    Returns:
        指定目录下所有文件夹的全路径名的列表
    """
    folder2fullpaths = {}
    for root, dirs, _ in os.walk(root_dir):
        for dir in dirs:
            full_path = os.path.join(root, dir)
            folder2fullpaths[dir] = full_path
    return folder2fullpaths


def write_dict_to_excel(data_dict, excel_file, sheet_name):
    """
    将字典数据写入 Excel

    Args:
        data_dict: 字典数据
        excel_file: Excel 文件名
        sheet_name: 工作表名称
    """
    try:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook[sheet_name]
        worksheet.delete_rows(1, worksheet.max_row)  # 清空工作表内容
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

    for fdrname, fdrnamefullpath in data_dict.items():
        worksheet.append([fdrname, fdrnamefullpath])

    workbook.save(excel_file)

basepath = "F:\文档\钢结构-工程测量-加固改造--智能建造test"
destpath = "F:\资料分类"
exl_name="资料分类.xlsx"
sheet_name='搜索词对应分类文件夹名'

print('当前工作目录 :')
print(os.getcwd(), '\n')

#all_folderstt = get_full_folder_paths(basepath)

wb = openpyxl.load_workbook('C:\\Users\\jimee\\Desktop\\方案清单目录V3-1.xlsx')
print('wb 类型 :')
print(type(wb), '\n')

# 选择工作表
sheet = wb['房建类工程方案清单']  
print('表名 - ' + sheet.title, '\n')

# 获取工作表中所有合并单元格的范围
merged_ranges = sheet.merged_cells.ranges

#1.目录列表
dataClassifyxlsx = os.path.join(destpath, exl_name)
wb_classify = openpyxl.load_workbook(dataClassifyxlsx)
sheet2 = wb_classify[sheet_name]  
col = 1
col2 = 2
secondname2secondlevel = {}
for row in range(1, 99 + 1):
    cell = sheet2.cell(row=row, column=col)
    cell2 = sheet2.cell(row=row, column=col2)
    secondname2secondlevel[cell2.value] = cell.value

#2.获取搜索词 每一个单以搜索词对应一个二级目录名存在key2secondnameDic里
data = []
key2secondnameDic = {} #(单搜索词：二级名)；
#2.1合并单元格
for merged_cell_range in sheet.merged_cells.ranges:
    min_col, min_row, max_col, max_row = merged_cell_range.bounds
    #选取第4列的合并单元格
    search_termlist = []
    if min_col == 4:
        start_cell = sheet.cell(row=min_row, column=min_col)
        search_term = start_cell.value
        #data.append(search_term)
        search_termlist = [item.strip() for item in search_term.split('、')]
        #print(search_termlist)
        data.append(search_termlist)

        # 二级文件夹分合并单元格和单元格两种情况处理
        #合并单元格
        foldernameList = []
        for merged_cell_range in sheet.merged_cells.ranges:
            min_colT, min_rowT, max_colT, max_rowT = merged_cell_range.bounds
            if min_colT == 3:
                if (min_rowT >= min_row) and (min_rowT <= max_row):
                    start_cellT = sheet.cell(row=min_rowT, column=min_colT)
                    folder_nameT = start_cellT.value
                    foldernameList.append(folder_nameT)
        
        #单个单元格
        colT = 3
        for row in range(1, 172 + 1):
            cell = sheet.cell(row=row, column=colT)
            if not is_merged_cell(cell, merged_ranges):#不在合并单元格里边
                if(row >= min_row) and (row <= max_row):
                    folder_nameT = cell.value
                    foldernameList.append(folder_nameT)

        #单个搜索词匹配一个二级目录名
        for term in search_termlist:
            bfindfolder = False
            for foldername in foldernameList:
                if term in foldername:
                    bfindfolder = True
                    key2secondnameDic[term] = foldername
                    #if(len(foldernameList) != 1):
                    break
            if not bfindfolder:
                if len(foldernameList) > 0:
                    key2secondnameDic[term] = foldernameList[0]

#2.2第4列的单个单元格
col = 4
for row in range(2, 172 + 1):
    cell = sheet.cell(row=row, column=col) 
    search_termlist = []
    #2.1排除在合并单元格里的
    if not is_merged_cell(cell, merged_ranges):
        search_term = cell.value
        #data.append(search_term)
        search_termlist = [item.strip() for item in search_term.split('、')]
        data.append(search_termlist)

        # 二级文件夹分合并单元格和单元格两种情况处理
        foldernameList = []
        """以下情况不存在
        #合并单元格
       
        for merged_cell_range in sheet.merged_cells.ranges:
            min_colT, min_rowT, max_colT, max_rowT = merged_cell_range.bounds
            if min_colT == 3:
                if (min_rowT == row):
                    start_cellT = sheet.cell(row=min_rowT, column=min_colT)
                    folder_nameT = start_cellT.value
                    foldernameList.append(folder_nameT)
        """
        #单个单元格
        colT = 3
        for rowT in range(1, 172 + 1):
            cell = sheet.cell(row=rowT, column=colT)
            if not is_merged_cell(cell, merged_ranges):#不在合并单元格里边
                if(rowT == row):
                    folder_nameT = cell.value
                    foldernameList.append(folder_nameT)
                    
        #单个搜索词匹配一个二级目录名
        for term in search_termlist:
            bfindfolder = False
            for foldername in foldernameList:
                if term in foldername:
                    bfindfolder = True
                    key2secondnameDic[term] = foldername
                    #if(len(foldernameList) != 1):
                    break
            if not bfindfolder:
                if len(foldernameList) > 0:
                    key2secondnameDic[term] = foldernameList[0]



excel_file = os.path.join(destpath, exl_name)
write_dict_to_excel(key2secondnameDic, excel_file, sheet_name)